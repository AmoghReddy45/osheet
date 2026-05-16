"""Generate a realistic dummy financial model .xlsx for benchmarking."""
import io
import openpyxl
from openpyxl.styles import PatternFill, Font


def make_financial_model() -> bytes:
    wb = openpyxl.Workbook()

    # Assumptions sheet
    ws_a = wb.active
    ws_a.title = "Assumptions"
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    assumptions = [
        ("Growth Rate", 0.15),
        ("Churn Rate", 0.04),
        ("Gross Margin", 0.72),
        ("COGS %", 0.28),
        ("Sales Headcount", 12),
        ("ACV", 24000),
        ("NRR", 1.08),
    ]
    ws_a["A1"] = "Assumption"
    ws_a["B1"] = "Value"
    ws_a["A1"].font = Font(bold=True)
    ws_a["B1"].font = Font(bold=True)
    for i, (name, val) in enumerate(assumptions, start=2):
        ws_a[f"A{i}"] = name
        ws_a[f"B{i}"] = val
        ws_a[f"B{i}"].fill = yellow

    # Revenue sheet
    ws_r = wb.create_sheet("Revenue")
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    headers = ["Month", "New ARR", "Churned ARR", "Net ARR", "Cumulative ARR", "Gross Profit"]
    for col, h in enumerate(headers, 1):
        ws_r.cell(row=1, column=col, value=h).font = Font(bold=True)

    base_arr = 100000
    for i, month in enumerate(months, start=2):
        ws_r[f"A{i}"] = month
        ws_r[f"B{i}"] = f"=ROUND({base_arr}*(1+Assumptions!$B$1)^{i-2},0)"
        ws_r[f"C{i}"] = f"=ROUND(E{i-1}*Assumptions!$B$2,0)" if i > 2 else 0
        ws_r[f"D{i}"] = f"=B{i}-C{i}"
        ws_r[f"E{i}"] = f"=IF({i}=2,B{i},E{i-1}+D{i})"
        ws_r[f"F{i}"] = f"=E{i}*Assumptions!$B$3"

    # Totals row
    ws_r["A14"] = "Total"
    ws_r["B14"] = "=SUM(B2:B13)"
    ws_r["D14"] = "=SUM(D2:D13)"
    ws_r["F14"] = "=SUM(F2:F13)"

    # Summary sheet
    ws_s = wb.create_sheet("Summary")
    ws_s["A1"] = "Metric"
    ws_s["B1"] = "Value"
    ws_s["A1"].font = Font(bold=True)
    ws_s["B1"].font = Font(bold=True)
    metrics = [
        ("Total ARR", "=Revenue!E13"),
        ("Total Gross Profit", "=Revenue!F14"),
        ("Net ARR Added", "=Revenue!D14"),
        ("Gross Margin %", "=Assumptions!B3"),
        ("Churn Rate", "=Assumptions!B2"),
        ("Growth Rate", "=Assumptions!B1"),
    ]
    for i, (name, formula) in enumerate(metrics, start=2):
        ws_s[f"A{i}"] = name
        ws_s[f"B{i}"] = formula

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    import os
    os.makedirs("benchmarks", exist_ok=True)
    data = make_financial_model()
    out_path = "benchmarks/dummy_financial_model.xlsx"
    with open(out_path, "wb") as f:
        f.write(data)
    print(f"Written {len(data):,} bytes → {out_path}")
