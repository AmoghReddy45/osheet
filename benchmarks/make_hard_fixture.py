# benchmarks/make_hard_fixture.py
"""
Generate hard benchmark fixture: 15-sheet SaaS model, ~300 cells, no yellow fills.
Role classifier must work purely from formula graph structure.
"""
import io
import openpyxl
from openpyxl.styles import Font

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# 20 assumptions in Config sheet — no yellow fills
CONFIG = [
    ("arr_start",           1_000_000),
    ("growth_rate_monthly", 0.025),
    ("churn_rate_monthly",  0.015),
    ("gross_margin",        0.72),
    ("sales_headcount",     12),
    ("eng_headcount",       20),
    ("gna_headcount",       8),
    ("avg_salary",          120_000),
    ("payroll_tax_rate",    0.15),
    ("benefits_rate",       0.12),
    ("acv",                 24_000),
    ("nrr",                 1.08),
    ("cac_ratio",           1.2),
    ("magic_number",        0.8),
    ("burn_multiple",       1.5),
    ("runway_months",       18),
    ("cogs_pct",            0.28),
    ("rd_pct",              0.35),
    ("sm_pct",              0.40),
    ("gna_pct",             0.15),
]


def _build_config(ws) -> None:
    ws["A1"] = "name"
    ws["B1"] = "value"
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    for i, (name, val) in enumerate(CONFIG, start=2):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = val
        # NO yellow fill — tests role classifier without color hints


def _config_ref(row: int) -> str:
    """Return absolute Config cell ref, e.g. Config!$B$2."""
    return f"Config!$B${row}"


def _build_monthly(ws, month_idx: int) -> None:
    """
    Build a monthly P&L. month_idx is 0-based.
    CONFIG row mapping (1-indexed header, 2-indexed first data row):
      B2=arr_start, B3=growth_rate_monthly, B4=churn_rate_monthly,
      B5=gross_margin, B6=sales_hc, B7=eng_hc, B8=gna_hc, B9=avg_salary,
      B10=payroll_tax_rate, B11=benefits_rate, B18=cogs_pct, B19=rd_pct, B20=sm_pct, B21=gna_pct
    """
    ws["A1"] = "Metric"
    ws["B1"] = "Value"
    ws["A1"].font = Font(bold=True)

    n = month_idx  # months elapsed (0 for Jan)

    # ARR = arr_start * (1+growth)^n * (1-churn)^n
    ws["A2"] = "ARR"
    ws["B2"] = f"={_config_ref(2)}*(1+{_config_ref(3)})^{n}*(1-{_config_ref(4)})^{n}"

    # New ARR this month
    ws["A3"] = "New_ARR"
    ws["B3"] = f"={_config_ref(2)}*{_config_ref(3)}*(1+{_config_ref(3)})^{n}"

    # Churned ARR
    ws["A4"] = "Churned_ARR"
    ws["B4"] = f"=B2*{_config_ref(4)}"

    # Net ARR
    ws["A5"] = "Net_ARR"
    ws["B5"] = "=B3-B4"

    # Monthly Revenue = ARR / 12
    ws["A6"] = "Revenue"
    ws["B6"] = "=B2/12"

    # COGS
    ws["A7"] = "COGS"
    ws["B7"] = f"=B6*{_config_ref(18)}"

    # Gross Profit
    ws["A8"] = "Gross_Profit"
    ws["B8"] = "=B6-B7"

    # R&D
    ws["A9"] = "RD"
    ws["B9"] = f"=B6*{_config_ref(19)}"

    # Sales & Marketing
    ws["A10"] = "Sales_Marketing"
    ws["B10"] = f"=B6*{_config_ref(20)}"

    # G&A
    ws["A11"] = "GnA"
    ws["B11"] = f"=B6*{_config_ref(21)}"

    # Total OpEx
    ws["A12"] = "Total_OpEx"
    ws["B12"] = "=B9+B10+B11"

    # EBITDA
    ws["A13"] = "EBITDA"
    ws["B13"] = "=B8-B12"

    # HC cost: (sales_hc + eng_hc + gna_hc) * avg_salary / 12 * (1 + payroll_tax + benefits)
    ws["A14"] = "HC_Cost"
    ws["B14"] = f"=({_config_ref(6)}+{_config_ref(7)}+{_config_ref(8)})*{_config_ref(9)}/12*(1+{_config_ref(10)}+{_config_ref(11)})"

    # EBITDA after HC
    ws["A15"] = "EBITDA_after_HC"
    ws["B15"] = "=B13-B14"


def _build_annual(ws) -> None:
    ws["A1"] = "Metric"
    ws["B1"] = "Annual"
    ws["A1"].font = Font(bold=True)

    def month_sum(row: int) -> str:
        return "+".join(f"{m}!B{row}" for m in MONTHS)

    metrics = [
        ("ARR_Dec",            "=Dec!B2"),
        ("Total_Revenue",      f"={month_sum(6)}"),
        ("Total_COGS",         f"={month_sum(7)}"),
        ("Total_Gross_Profit", f"={month_sum(8)}"),
        ("Total_RD",           f"={month_sum(9)}"),
        ("Total_SM",           f"={month_sum(10)}"),
        ("Total_GnA",          f"={month_sum(11)}"),
        ("Total_OpEx",         f"={month_sum(12)}"),
        ("Total_EBITDA",       f"={month_sum(13)}"),
        ("Total_HC_Cost",      f"={month_sum(14)}"),
        ("Net_EBITDA",         f"={month_sum(15)}"),
    ]
    for i, (name, formula) in enumerate(metrics, start=2):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = formula


def _build_kpis(ws) -> None:
    ws["A1"] = "KPI"
    ws["B1"] = "Value"
    ws["A1"].font = Font(bold=True)

    kpis = [
        ("Gross_Margin_Pct", "=Annual!B4/Annual!B3"),
        ("ARR_Growth_Pct",   f"=(Annual!B2-{_config_ref(2)})/{_config_ref(2)}"),
        ("Burn_Multiple",    "=ABS(Annual!B9)/Annual!B2"),
        ("NRR",              f"={_config_ref(12)}"),
        ("Magic_Number",     f"={_config_ref(14)}"),
        ("Rule_of_40",       "=KPIs!B2*100+Annual!B2/1000000"),
    ]
    for i, (name, formula) in enumerate(kpis, start=2):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = formula


def make_hard_model() -> bytes:
    wb = openpyxl.Workbook()

    ws_config = wb.active
    ws_config.title = "Config"
    _build_config(ws_config)

    for i, month in enumerate(MONTHS):
        _build_monthly(wb.create_sheet(month), i)

    _build_annual(wb.create_sheet("Annual"))
    _build_kpis(wb.create_sheet("KPIs"))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":
    import os
    data = make_hard_model()
    out = "benchmarks/hard_financial_model.xlsx"
    with open(out, "wb") as f:
        f.write(data)
    sheets = 1 + len(MONTHS) + 2  # Config + 12 monthly + Annual + KPIs
    print(f"Written {len(data):,} bytes → {out}")
    print(f"Sheets: {sheets} (Config + 12 monthly + Annual + KPIs)")
    print(f"Config assumptions: {len(CONFIG)} (no yellow fills)")
