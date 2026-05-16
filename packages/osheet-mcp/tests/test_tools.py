import io
import openpyxl
import pytest
from osheet_mcp.server import (
    get_workbook_summary,
    get_assumptions,
    get_outputs,
    trace_cell,
    find_cells,
    propose_patch_tool,
    clear_cache,
)


@pytest.fixture
def xlsx_path(tmp_path) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Model"
    ws["A1"] = "Growth Rate"
    ws["B1"] = 0.15
    ws["A2"] = "Base Revenue"
    ws["B2"] = 500_000
    ws["A3"] = "Year 1 Revenue"
    ws["B3"] = "=B2*(1+B1)"
    ws["A4"] = "Year 2 Revenue"
    ws["B4"] = "=B3*(1+B1)"
    ws["A5"] = "Total"
    ws["B5"] = "=SUM(B3:B4)"
    path = str(tmp_path / "model.xlsx")
    wb.save(path)
    return path


def setup_function():
    clear_cache()


def test_get_workbook_summary_returns_string(xlsx_path):
    result = get_workbook_summary(xlsx_path)
    assert isinstance(result, str)
    assert "sheet" in result.lower()


def test_get_assumptions_returns_list(xlsx_path):
    result = get_assumptions(xlsx_path)
    assert isinstance(result, list)
    assert len(result) >= 1


def test_get_assumptions_have_required_fields(xlsx_path):
    result = get_assumptions(xlsx_path)
    for item in result:
        assert "stable_id" in item
        assert "value" in item
        assert "sheet" in item


def test_get_outputs_returns_list(xlsx_path):
    result = get_outputs(xlsx_path)
    assert isinstance(result, list)


def test_trace_cell_returns_upstream_and_downstream(xlsx_path):
    assumptions = get_assumptions(xlsx_path)
    if not assumptions:
        pytest.skip("no assumptions")
    result = trace_cell(xlsx_path, assumptions[0]["stable_id"])
    assert "upstream" in result
    assert "downstream" in result
    assert isinstance(result["upstream"], list)
    assert isinstance(result["downstream"], list)


def test_find_cells_returns_list(xlsx_path):
    result = find_cells(xlsx_path, "growth")
    assert isinstance(result, list)


def test_propose_patch_returns_diff(xlsx_path):
    assumptions = get_assumptions(xlsx_path)
    if not assumptions:
        pytest.skip("no assumptions")
    a = next((x for x in assumptions if isinstance(x["value"], (int, float))), None)
    if not a:
        pytest.skip("no numeric assumption")
    result = propose_patch_tool(xlsx_path, a["stable_id"], 0.30)
    assert "cell_id" in result
    assert "new_value" in result
    assert "diff" in result
    assert "computed_values" in result


def test_propose_patch_computed_values_dict(xlsx_path):
    assumptions = get_assumptions(xlsx_path)
    if not assumptions:
        pytest.skip("no assumptions")
    a = next((x for x in assumptions if isinstance(x["value"], (int, float))), None)
    if not a:
        pytest.skip("no numeric assumption")
    result = propose_patch_tool(xlsx_path, a["stable_id"], 0.30)
    assert isinstance(result["computed_values"], dict)
