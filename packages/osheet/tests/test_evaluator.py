# packages/osheet/tests/test_evaluator.py
import datetime as dt
import io
import openpyxl
import pytest
from osheet.parser import parse_xlsx
from osheet.analyzer import run_all
from osheet.evaluator import evaluate_patch


@pytest.fixture
def two_sheet_bytes() -> bytes:
    """Config (assumptions) + Revenue (formula cells)."""
    wb = openpyxl.Workbook()
    ws_c = wb.active
    ws_c.title = "Config"
    ws_c["A1"] = "growth_rate"
    ws_c["B1"] = 0.10          # assumption — 10% growth
    ws_c["A2"] = "base_arr"
    ws_c["B2"] = 100_000       # assumption — 100k base

    ws_r = wb.create_sheet("Revenue")
    ws_r["A1"] = "Year 1"
    ws_r["B1"] = "=Config!B2*(1+Config!B1)"   # = 110,000
    ws_r["A2"] = "Year 2"
    ws_r["B2"] = "=B1*(1+Config!B1)"           # = 121,000
    ws_r["A3"] = "Total"
    ws_r["B3"] = "=SUM(B1:B2)"                 # = 231,000

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_evaluate_baseline_values(two_sheet_bytes):
    workbook = parse_xlsx(two_sheet_bytes)
    run_all(workbook)
    # No patch — should return current values
    result = evaluate_patch({}, workbook)
    year1 = next(c for c in workbook.all_cells if c.formula and "Config!B2" in c.formula and "Config!B1" in c.formula)
    assert abs(result[year1.stable_id] - 110_000) < 1


def test_evaluate_patch_growth_rate(two_sheet_bytes):
    workbook = parse_xlsx(two_sheet_bytes)
    run_all(workbook)

    growth = next(c for c in workbook.all_cells if c.value == 0.10)
    result = evaluate_patch({growth.stable_id: 0.20}, workbook)

    # Year 1 = 100000 * 1.20 = 120,000
    year1 = next(c for c in workbook.all_cells if c.formula and "Config!B2" in c.formula)
    assert abs(result[year1.stable_id] - 120_000) < 1


def test_evaluate_patch_base_arr(two_sheet_bytes):
    workbook = parse_xlsx(two_sheet_bytes)
    run_all(workbook)

    base = next(c for c in workbook.all_cells if c.value == 100_000)
    result = evaluate_patch({base.stable_id: 200_000}, workbook)

    year1 = next(c for c in workbook.all_cells if c.formula and "Config!B2" in c.formula)
    assert abs(result[year1.stable_id] - 220_000) < 1


def test_evaluate_sum_propagates(two_sheet_bytes):
    workbook = parse_xlsx(two_sheet_bytes)
    run_all(workbook)

    growth = next(c for c in workbook.all_cells if c.value == 0.10)
    result = evaluate_patch({growth.stable_id: 0.20}, workbook)

    # Total = Year1 + Year2 = 120000 + 144000 = 264000
    total = next(c for c in workbook.all_cells if c.formula and "SUM" in c.formula)
    assert abs(result[total.stable_id] - 264_000) < 1


def test_evaluate_bad_formula_returns_none(two_sheet_bytes):
    """Cells with unparseable formulas should return None, not the stale value."""
    workbook = parse_xlsx(two_sheet_bytes)
    run_all(workbook)
    # Find a real formula cell and inject a bad formula
    formula_cell = next(c for c in workbook.all_cells if c.formula)
    original_formula = formula_cell.formula
    formula_cell.formula = "=)BROKEN("
    formula_cell.value = 99999  # stale value before bad formula
    result = evaluate_patch({}, workbook)
    # Restore
    formula_cell.formula = original_formula
    # Should return None for failed formula, not the stale 99999
    assert result[formula_cell.stable_id] is None


@pytest.fixture
def self_ref_bytes() -> bytes:
    """A1 = A1 (self-reference, seeds to 0, converges to 0)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "=A1"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


@pytest.fixture
def two_cell_cycle_bytes() -> bytes:
    """
    A1 = 1000 (hardcoded)
    B1 = (A1 + C1) / 2     depends on C1
    C1 = B1 * 0.1           depends on B1  --> circular B1<->C1
    Fixed point: B1 = 1000/1.9 ≈ 526.3158, C1 = B1*0.1 ≈ 52.6316
    """
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Sheet1"
    ws["A1"] = 1000
    ws["B1"] = "=(A1+C1)/2"
    ws["C1"] = "=B1*0.1"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


@pytest.fixture
def interest_chain_bytes() -> bytes:
    """
    Classic 3-statement interest circularity (simplified):
    A1 = 10_000   (EBITDA, hardcoded)
    A2 = 0.05     (interest rate, hardcoded)
    A3 = A1 - A4  (EBT = EBITDA - interest expense; depends on A4 circular)
    A4 = A5 * A2  (interest expense = avg_debt * rate; depends on A5 circular)
    A5 = A3 * 2   (avg_debt = EBT * 2, simplified; depends on A3 circular)
    Fixed point:
      A4 = A5*0.05 = A3*2*0.05 = A3*0.1
      A3 = A1 - A4 = 10000 - A3*0.1
      1.1*A3 = 10000 -> A3 = 9090.909, A4 = 909.09, A5 = 18181.8
    """
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Sheet1"
    ws["A1"] = 10_000
    ws["A2"] = 0.05
    ws["A3"] = "=A1-A4"
    ws["A4"] = "=A5*A2"
    ws["A5"] = "=A3*2"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


@pytest.fixture
def diverging_cycle_bytes() -> bytes:
    """
    B1 = C1 * 3   (spectral radius = 9, diverges rapidly)
    C1 = B1 * 3
    Should not raise; should return (without hanging) even if not converged.
    """
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Sheet1"
    ws["B1"] = "=C1*3"
    ws["C1"] = "=B1*3"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


@pytest.fixture
def circular_with_downstream_bytes() -> bytes:
    """
    A1 = 100 (hardcoded)
    B1 = (A1 + C1) / 2      circular with C1
    C1 = B1 * 0.1            circular with B1
    D1 = B1 + 500            downstream of circular, non-circular
    Fixed point: B1 ≈ 52.6316, so D1 ≈ 552.6316
    """
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Sheet1"
    ws["A1"] = 100
    ws["B1"] = "=(A1+C1)/2"
    ws["C1"] = "=B1*0.1"
    ws["D1"] = "=B1+500"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def test_self_reference_converges(self_ref_bytes):
    """Self-referential cell seeds to 0 and stays at 0."""
    workbook = parse_xlsx(self_ref_bytes)
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    a1 = next(c for c in workbook.all_cells if c.formula)
    assert result[a1.stable_id] == 0


def test_two_cell_cycle_converges(two_cell_cycle_bytes):
    """B1 and C1 in a cycle converge to the analytic fixed point."""
    workbook = parse_xlsx(two_cell_cycle_bytes)
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    b1 = next(c for c in workbook.all_cells if c.formula and "A1" in (c.formula or "") and "C1" in (c.formula or ""))
    c1 = next(c for c in workbook.all_cells if c.formula and "B1" in (c.formula or "") and "0.1" in (c.formula or ""))
    assert abs(result[b1.stable_id] - (1000 / 1.9)) < 0.1
    assert abs(result[c1.stable_id] - (100 / 1.9)) < 0.1


def test_interest_chain_converges(interest_chain_bytes):
    """3-cell interest chain converges to analytic fixed point."""
    workbook = parse_xlsx(interest_chain_bytes)
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    # A3 (EBT) should be ≈ 9090.91
    a3 = next(c for c in workbook.all_cells if c.formula and "A4" in (c.formula or "") and "A1" in (c.formula or ""))
    assert abs(result[a3.stable_id] - (10_000 / 1.1)) < 1.0


def test_diverging_cycle_does_not_crash(diverging_cycle_bytes):
    """Diverging circular chain completes without exception."""
    workbook = parse_xlsx(diverging_cycle_bytes)
    run_all(workbook)
    # Must not raise; return value may be large numbers or None
    result = evaluate_patch({}, workbook)
    assert isinstance(result, dict)


def test_downstream_of_circular_uses_converged_value(circular_with_downstream_bytes):
    """D1 = B1 + 500 where B1 is circular; D1 should use converged B1."""
    workbook = parse_xlsx(circular_with_downstream_bytes)
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    d1 = next(c for c in workbook.all_cells if c.formula and "500" in (c.formula or ""))
    b1 = next(c for c in workbook.all_cells if c.formula and "A1" in (c.formula or "") and "C1" in (c.formula or ""))
    b1_val = result[b1.stable_id]
    d1_val = result[d1.stable_id]
    assert b1_val is not None
    assert d1_val is not None
    assert abs(d1_val - (b1_val + 500)) < 0.1


# --- OFFSET-as-range-endpoint tests -----------------------------------------


def _single_sheet_bytes(cells: dict, title: str = "Sheet1") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    for addr, val in cells.items():
        ws[addr] = val
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_offset_simple_negative_col():
    """E1 = SUM(A1:OFFSET(A1,0,2)) → A1:C1 = 1+2+3 = 6."""
    data = {
        "A1": 1, "B1": 2, "C1": 3, "D1": 4,
        "E1": "=SUM(A1:OFFSET(A1,0,2))",
    }
    workbook = parse_xlsx(_single_sheet_bytes(data))
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    e1 = next(c for c in workbook.all_cells if c.formula and "OFFSET" in c.formula)
    assert result[e1.stable_id] == 6


def test_offset_inside_average_with_min():
    """=AVERAGE(L322:OFFSET(L322,0,MIN(2,L5-1))) → L322:N322 → avg of 1,2,3 = 2.0"""
    data = {
        "L322": 1, "M322": 2, "N322": 3, "O322": 4, "P322": 5,
        "L5": 3,
        "A1": "=AVERAGE(L322:OFFSET(L322,0,MIN(2,L5-1)))",
    }
    workbook = parse_xlsx(_single_sheet_bytes(data))
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    a1 = next(c for c in workbook.all_cells if c.formula and "AVERAGE" in c.formula)
    assert abs(result[a1.stable_id] - 2.0) < 1e-9


def test_offset_with_bare_cell_ref():
    """=AVERAGE(L322:OFFSET(L322,0,L5)) where L5=2 → L322:N322 → avg of 10,20,30 = 20.0"""
    data = {
        "L322": 10, "M322": 20, "N322": 30,
        "L5": 2,
        "A1": "=AVERAGE(L322:OFFSET(L322,0,L5))",
    }
    workbook = parse_xlsx(_single_sheet_bytes(data))
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    a1 = next(c for c in workbook.all_cells if c.formula and "AVERAGE" in c.formula)
    assert abs(result[a1.stable_id] - 20.0) < 1e-9


def test_offset_out_of_range_returns_none():
    """=OFFSET(A1,0,-5) → new col would be -4 → caller returns None."""
    data = {
        "A1": 1,
        "B1": "=OFFSET(A1,0,-5)",
    }
    workbook = parse_xlsx(_single_sheet_bytes(data))
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    b1 = next(c for c in workbook.all_cells if c.formula)
    assert result[b1.stable_id] is None


def test_offset_with_sheet_name():
    """Cross-sheet OFFSET reference resolves correctly."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Data"
    ws1["A1"] = 10
    ws1["B1"] = 20
    ws1["C1"] = 30
    ws2 = wb.create_sheet("Calc")
    ws2["A1"] = "=SUM(Data!A1:OFFSET(Data!A1,0,2))"
    buf = io.BytesIO()
    wb.save(buf)
    workbook = parse_xlsx(buf.getvalue())
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    calc = next(c for c in workbook.all_cells if c.formula and "SUM" in c.formula)
    assert result[calc.stable_id] == 60


# --- IF short-circuit (skip unresolvable OFFSET in dead branch) -------------


def test_if_offset_false_branch_short_circuits():
    """IF condition false -> OFFSET in true branch is skipped -> 99."""
    data = {
        "A1": 2,
        "B1": 7,
        "C1": "=IF(A1>=5, OFFSET(B1,0,-3), 99)",
    }
    workbook = parse_xlsx(_single_sheet_bytes(data))
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    c1 = next(c for c in workbook.all_cells if c.formula and "IF" in (c.formula or ""))
    assert result[c1.stable_id] == 99


def test_if_offset_true_branch_evaluates():
    """IF condition true with in-range OFFSET in active branch evaluates correctly."""
    data = {
        "A1": 10,
        "C1": 7,
        "D1": 99,
        "E1": "=IF(A1>=5, OFFSET(D1,0,-1), 99)",
    }
    workbook = parse_xlsx(_single_sheet_bytes(data))
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    e1 = next(c for c in workbook.all_cells if c.formula and "IF" in (c.formula or ""))
    assert result[e1.stable_id] == 7


def test_if_offset_in_else_branch():
    """IF condition true with unresolvable OFFSET in else branch -> 7."""
    data = {
        "A1": 10,
        "B1": 5,
        "C1": "=IF(A1>=5, 7, OFFSET(B1,0,-3))",
    }
    workbook = parse_xlsx(_single_sheet_bytes(data))
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    c1 = next(c for c in workbook.all_cells if c.formula and "IF" in (c.formula or ""))
    assert result[c1.stable_id] == 7


def test_nested_if_offset_short_circuits():
    """Nested IFs: outer true, inner never touched (contains unresolvable OFFSET)."""
    data = {
        "A1": 10,
        "A2": 2,
        "B1": 3,
        "C1": "=IF(A1>=5, 42, IF(A2>=5, 99, OFFSET(B1,0,-9)))",
    }
    workbook = parse_xlsx(_single_sheet_bytes(data))
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    c1 = next(c for c in workbook.all_cells if c.formula and "IF" in (c.formula or ""))
    assert result[c1.stable_id] == 42


def test_nested_if_offset_short_circuits_descends():
    """Nested IFs: outer false, inner false, OFFSET active and unresolvable -> None."""
    data = {
        "A1": 1,
        "A2": 2,
        "B1": 3,
        "C1": "=IF(A1>=5, 42, IF(A2>=5, 99, OFFSET(B1,0,-9)))",
    }
    workbook = parse_xlsx(_single_sheet_bytes(data))
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    c1 = next(c for c in workbook.all_cells if c.formula and "IF" in (c.formula or ""))
    # OFFSET is in the active branch and unresolvable -> None
    assert result[c1.stable_id] is None


def test_if_offset_reachable_and_in_range():
    """IF true, OFFSET in active branch is in range and resolves."""
    data = {
        "A1": 10,
        "B1": 100, "C1": 200, "D1": 300, "E1": 400,
        "F1": "=IF(A1>=5, OFFSET(E1,0,-2), 0)",
    }
    workbook = parse_xlsx(_single_sheet_bytes(data))
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    f1 = next(c for c in workbook.all_cells if c.formula and "IF" in (c.formula or ""))
    # OFFSET(E1, 0, -2) -> C1 = 200
    assert result[f1.stable_id] == 200


# --- Structured-reference (Table[[#This Row],[Col]]) tests ------------------


def test_structured_ref_this_row_resolves():
    """=MyTbl[[#This Row],[B_col]]+1 at C2 -> B2+1 = 11."""
    from openpyxl.worksheet.table import Table, TableColumn
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws['A1'], ws['B1'] = 'A_col', 'B_col'
    ws['A2'], ws['B2'] = 1, 10
    ws['A3'], ws['B3'] = 2, 20
    ws['C2'] = "=MyTbl[[#This Row],[B_col]]+1"
    tbl = Table(displayName='MyTbl', name='MyTbl', ref='A1:B3', tableColumns=[
        TableColumn(id=1, name='A_col'),
        TableColumn(id=2, name='B_col'),
    ])
    ws.add_table(tbl)
    buf = io.BytesIO()
    wb.save(buf)
    workbook = parse_xlsx(buf.getvalue())
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    c2 = next(c for c in workbook.all_cells if c.formula and "MyTbl" in c.formula)
    assert result[c2.stable_id] == 11


def test_structured_ref_unknown_table_returns_none():
    """Reference to undefined table -> result is None, no crash."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws['A1'] = 5
    ws['B1'] = "=NoSuchTable[[#This Row],[NoCol]]+1"
    buf = io.BytesIO()
    wb.save(buf)
    workbook = parse_xlsx(buf.getvalue())
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    b1 = next(c for c in workbook.all_cells if c.formula)
    assert result[b1.stable_id] is None


def test_structured_ref_cross_sheet():
    """Table on Sheet1, formula on Sheet2 references the table."""
    from openpyxl.worksheet.table import Table, TableColumn
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Data Sheet"
    ws1['A1'], ws1['B1'] = 'A_col', 'B_col'
    ws1['A2'], ws1['B2'] = 1, 100
    ws1['A3'], ws1['B3'] = 2, 200
    tbl = Table(displayName='CrossTbl', name='CrossTbl', ref='A1:B3', tableColumns=[
        TableColumn(id=1, name='A_col'),
        TableColumn(id=2, name='B_col'),
    ])
    ws1.add_table(tbl)
    ws2 = wb.create_sheet("Calc")
    # Formula at C2 referencing #This Row -> row 2 of Data Sheet -> B2 = 100
    ws2['C2'] = "=CrossTbl[[#This Row],[B_col]]*2"
    buf = io.BytesIO()
    wb.save(buf)
    workbook = parse_xlsx(buf.getvalue())
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    c2 = next(c for c in workbook.all_cells if c.formula and "CrossTbl" in c.formula)
    assert result[c2.stable_id] == 200


# --- Whole-column / totals / headers structured-reference tests -------------


def _whole_col_workbook(formula: str, *, totals: bool = False) -> bytes:
    """Helper: 3 data rows (Code, Amount). Optionally with totals row."""
    from openpyxl.worksheet.table import Table, TableColumn
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws['A1'], ws['B1'] = 'Code', 'Amount'
    ws['A2'], ws['B2'] = 'X', 10
    ws['A3'], ws['B3'] = 'Y', 20
    ws['A4'], ws['B4'] = 'X', 30
    if totals:
        ws['A5'], ws['B5'] = 'Total', 60
        ref = 'A1:B5'
        tbl_kwargs = {'totalsRowCount': 1}
    else:
        ref = 'A1:B4'
        tbl_kwargs = {}
    ws['D1'] = formula
    tbl = Table(
        displayName='MyTbl', name='MyTbl', ref=ref,
        tableColumns=[
            TableColumn(id=1, name='Code'),
            TableColumn(id=2, name='Amount'),
        ],
        **tbl_kwargs,
    )
    ws.add_table(tbl)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_structured_ref_whole_column_in_sum():
    """=SUM(MyTbl[Amount]) over [10,20,30] -> 60."""
    workbook = parse_xlsx(_whole_col_workbook("=SUM(MyTbl[Amount])"))
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    d1 = next(c for c in workbook.all_cells if c.formula and "MyTbl" in c.formula)
    assert result[d1.stable_id] == 60


def test_structured_ref_whole_column_in_sumif():
    """=SUMIF(MyTbl[Code], "X", MyTbl[Amount]) -> 10+30 = 40."""
    workbook = parse_xlsx(_whole_col_workbook('=SUMIF(MyTbl[Code], "X", MyTbl[Amount])'))
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    d1 = next(c for c in workbook.all_cells if c.formula and "MyTbl" in c.formula)
    assert result[d1.stable_id] == 40


def test_structured_ref_totals_row_single_col():
    """=MyTbl[[#Totals],[Amount]] -> 60 (the totals row value)."""
    workbook = parse_xlsx(_whole_col_workbook("=MyTbl[[#Totals],[Amount]]", totals=True))
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    d1 = next(c for c in workbook.all_cells if c.formula and "MyTbl" in c.formula)
    assert result[d1.stable_id] == 60


def test_structured_ref_totals_excludes_totals_row_from_data():
    """=SUM(MyTbl[Amount]) on table with totals row sums data only (10+20+30=60)."""
    workbook = parse_xlsx(_whole_col_workbook("=SUM(MyTbl[Amount])", totals=True))
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    d1 = next(c for c in workbook.all_cells if c.formula and "MyTbl" in c.formula)
    # If totals row were included, sum would be 120 instead.
    assert result[d1.stable_id] == 60


def test_structured_ref_no_totals_row_totals_ref_is_none():
    """=MyTbl[[#Totals],[Amount]] without a totals row -> None (unrewritable)."""
    workbook = parse_xlsx(_whole_col_workbook("=MyTbl[[#Totals],[Amount]]", totals=False))
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    d1 = next(c for c in workbook.all_cells if c.formula and "MyTbl" in c.formula)
    assert result[d1.stable_id] is None


def test_structured_ref_column_range():
    """=SUM(MyTbl[[Jan]:[Mar]]) over 2D area -> sum of all values."""
    from openpyxl.worksheet.table import Table, TableColumn
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws['A1'], ws['B1'], ws['C1'] = 'Jan', 'Feb', 'Mar'
    ws['A2'], ws['B2'], ws['C2'] = 1, 2, 3
    ws['A3'], ws['B3'], ws['C3'] = 4, 5, 6
    ws['E1'] = "=SUM(MyTbl[[Jan]:[Mar]])"
    tbl = Table(displayName='MyTbl', name='MyTbl', ref='A1:C3', tableColumns=[
        TableColumn(id=1, name='Jan'),
        TableColumn(id=2, name='Feb'),
        TableColumn(id=3, name='Mar'),
    ])
    ws.add_table(tbl)
    buf = io.BytesIO()
    wb.save(buf)
    workbook = parse_xlsx(buf.getvalue())
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    e1 = next(c for c in workbook.all_cells if c.formula and "MyTbl" in c.formula)
    assert result[e1.stable_id] == 21  # 1+2+3+4+5+6


def test_structured_ref_headers_single_col():
    """=MyTbl[[#Headers],[Amount]] -> 'Amount' (the header label)."""
    workbook = parse_xlsx(_whole_col_workbook("=MyTbl[[#Headers],[Amount]]"))
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    d1 = next(c for c in workbook.all_cells if c.formula and "MyTbl" in c.formula)
    assert result[d1.stable_id] == "Amount"


def test_structured_ref_iferror_wraps_table_ref():
    """=IFERROR(SUM(MyTbl[Amount])/COUNTA(MyTbl[Amount]),0) -> mean = 60/3 = 20."""
    workbook = parse_xlsx(
        _whole_col_workbook("=IFERROR(SUM(MyTbl[Amount])/COUNTA(MyTbl[Amount]),0)")
    )
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    d1 = next(c for c in workbook.all_cells if c.formula and "MyTbl" in c.formula)
    assert abs(result[d1.stable_id] - 20.0) < 1e-9


def test_structured_ref_subtotal_109_whole_column():
    """=SUBTOTAL(109, MyTbl[TOTAL COST]) -> sum of the column."""
    from openpyxl.worksheet.table import Table, TableColumn
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws['A1'] = 'TOTAL COST'
    ws['A2'] = 100
    ws['A3'] = 200
    ws['A4'] = 300
    ws['C1'] = "=SUBTOTAL(109, MyTbl[TOTAL COST])"
    tbl = Table(displayName='MyTbl', name='MyTbl', ref='A1:A4', tableColumns=[
        TableColumn(id=1, name='TOTAL COST'),
    ])
    ws.add_table(tbl)
    buf = io.BytesIO()
    wb.save(buf)
    workbook = parse_xlsx(buf.getvalue())
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    c1 = next(c for c in workbook.all_cells if c.formula and "MyTbl" in c.formula)
    assert result[c1.stable_id] == 600


def test_structured_ref_bare_brackets_with_unknown_table_unchanged():
    """Bare [X] with no matching table -> None (no crash)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws['A1'] = 5
    ws['B1'] = "=NoSuchTbl[X]+1"
    buf = io.BytesIO()
    wb.save(buf)
    workbook = parse_xlsx(buf.getvalue())
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    b1 = next(c for c in workbook.all_cells if c.formula)
    assert result[b1.stable_id] is None


# --- Date pipeline tests (TODAY/DATE/YEAR/EDATE/EOMONTH/TEXT/arithmetic) -----


@pytest.fixture
def date_bytes() -> bytes:
    """Workbook with date inputs and a variety of date-formula outputs."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Config"
    ws['B3'] = dt.datetime(2020, 1, 1)
    ws['B4'] = dt.datetime(2019, 6, 1)
    ws['C3'] = "=YEAR(B3)"
    ws['C4'] = "=EDATE(B3, 12)"
    ws['C5'] = "=EOMONTH(B3, 0)"
    ws['C6'] = "=TEXT(B3, \"mmmm\")"
    ws['C7'] = "=B3 + 30"
    ws['C8'] = "=B3 > B4"
    ws['C9'] = "=DATE(2026, 5, 16)"
    ws['C10'] = "=TODAY()"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell_by_formula(workbook, needle: str):
    return next(c for c in workbook.all_cells if c.formula and needle in c.formula)


def test_year_of_date_cell(date_bytes):
    """=YEAR(B3) where B3 is 2020-01-01 -> 2020."""
    workbook = parse_xlsx(date_bytes)
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    cell = _cell_by_formula(workbook, "YEAR(B3)")
    assert int(result[cell.stable_id]) == 2020


def test_edate_one_year_forward(date_bytes):
    """=EDATE(B3, 12) where B3 = 2020-01-01 -> 2021-01-01 serial 44197."""
    workbook = parse_xlsx(date_bytes)
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    cell = _cell_by_formula(workbook, "EDATE")
    assert abs(float(result[cell.stable_id]) - 44197) <= 0.5


def test_eomonth_last_day(date_bytes):
    """=EOMONTH(B3, 0) where B3 = 2020-01-01 -> 2020-01-31 serial 43861."""
    workbook = parse_xlsx(date_bytes)
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    cell = _cell_by_formula(workbook, "EOMONTH")
    assert abs(float(result[cell.stable_id]) - 43861) <= 0.5


def test_date_function_returns_serial(date_bytes):
    """=DATE(2026, 5, 16) -> serial 46158."""
    workbook = parse_xlsx(date_bytes)
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    cell = _cell_by_formula(workbook, "DATE(2026")
    assert abs(float(result[cell.stable_id]) - 46158) <= 0.5


def test_today_returns_today_serial(date_bytes):
    """=TODAY() -> serial of today's date (within ±1 day for timezone)."""
    workbook = parse_xlsx(date_bytes)
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    cell = _cell_by_formula(workbook, "TODAY")
    expected = (dt.date.today() - dt.date(1899, 12, 30)).days
    assert abs(float(result[cell.stable_id]) - expected) <= 1


def test_text_month_name(date_bytes):
    """=TEXT(B3, "mmmm") where B3 = 2020-01-01 -> "January"."""
    workbook = parse_xlsx(date_bytes)
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    cell = _cell_by_formula(workbook, "TEXT")
    assert result[cell.stable_id] == "January"


def test_date_arithmetic_add_days(date_bytes):
    """=B3 + 30 where B3 = 2020-01-01 -> 2020-01-31 serial 43861."""
    workbook = parse_xlsx(date_bytes)
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    cell = _cell_by_formula(workbook, "B3 + 30")
    assert abs(float(result[cell.stable_id]) - 43861) <= 0.5


def test_date_comparison(date_bytes):
    """=B3 > B4 where B3 = 2020-01-01, B4 = 2019-06-01 -> True."""
    workbook = parse_xlsx(date_bytes)
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    cell = _cell_by_formula(workbook, "B3 > B4")
    assert bool(result[cell.stable_id]) is True


# --- Range resolver tests ---------------------------------------------------


def test_expand_range_bare():
    vm = {"S!A1": 1, "S!B1": 2, "S!C1": 3}
    from osheet.evaluator import _expand_range
    arr = _expand_range("A1:C1", "S", vm)
    assert list(arr[0]) == [1.0, 2.0, 3.0]


def test_expand_range_with_sheet_both_sides():
    vm = {"S!A1": 1, "S!B1": 2, "S!C1": 3}
    from osheet.evaluator import _expand_range
    arr = _expand_range("S!A1:S!C1", "OTHER", vm)
    assert list(arr[0]) == [1.0, 2.0, 3.0]


def test_expand_range_with_sheet_left_only():
    vm = {"S!A1": 1, "S!B1": 2, "S!C1": 3}
    from osheet.evaluator import _expand_range
    arr = _expand_range("S!A1:C1", "OTHER", vm)
    assert list(arr[0]) == [1.0, 2.0, 3.0]


def test_expand_range_dollar_anchored():
    vm = {"S!A1": 1, "S!B1": 2, "S!C1": 3}
    from osheet.evaluator import _expand_range
    arr = _expand_range("$A$1:$C$1", "S", vm)
    assert list(arr[0]) == [1.0, 2.0, 3.0]


def test_expand_range_mixed_quoting():
    vm = {"TBA!A1": 1, "TBA!B1": 2, "TBA!C1": 3}
    from osheet.evaluator import _expand_range
    arr = _expand_range("TBA!$A$1:'TBA'!C1", "OTHER", vm)
    assert list(arr[0]) == [1.0, 2.0, 3.0]


def test_expand_range_quoted_sheet_with_space():
    vm = {"MY SHEET!A1": 1, "MY SHEET!B1": 2, "MY SHEET!C1": 3}
    from osheet.evaluator import _expand_range
    arr = _expand_range("'My Sheet'!A1:'My Sheet'!C1", "X", vm)
    assert list(arr[0]) == [1.0, 2.0, 3.0]


def test_expand_range_formulas_lib_emitted_form():
    """formulas library emits ranges with surrounding parens and spaces around ':'"""
    vm = {"TBA!A1": 1, "TBA!B1": 2, "TBA!C1": 3}
    from osheet.evaluator import _expand_range
    arr = _expand_range("(TBA!A1: TBA!C1)", "X", vm)
    assert list(arr[0]) == [1.0, 2.0, 3.0]


def test_parse_refs_mixed_anchored_range():
    from osheet.analyzer.graph import _parse_refs
    refs = _parse_refs("=SUM(TBA!$D$10:'TBA'!I10)", "DEFAULT")
    cells = {(r[1], r[2]) for r in refs if r[0] == "TBA"}
    expected = {(c, 10) for c in range(4, 10)}  # D=4, I=9 inclusive
    assert cells == expected


def test_parse_refs_same_sheet_range():
    from osheet.analyzer.graph import _parse_refs
    refs = _parse_refs("=SUM(Sheet1!A1:Sheet1!C1)", "DEFAULT")
    cells_in_sheet1 = [(r[1], r[2]) for r in refs if r[0] == "Sheet1"]
    assert len(cells_in_sheet1) == 3  # A1, B1, C1


def test_to_float_comma_separated():
    from osheet.evaluator import _to_float
    assert _to_float("5,661") == 5661.0
    assert _to_float("1,234,567") == 1234567.0


def test_to_float_accounting_negative():
    from osheet.evaluator import _to_float
    assert _to_float("(2,032)") == -2032.0
    assert _to_float("(100)") == -100.0


def test_to_float_currency_prefix():
    from osheet.evaluator import _to_float
    assert _to_float("$5,661") == 5661.0


def test_to_float_percent():
    from osheet.evaluator import _to_float
    assert _to_float("50%") == 0.5
    assert _to_float("12.5%") == 0.125


def test_to_float_whitespace():
    from osheet.evaluator import _to_float
    assert _to_float("  5,661  ") == 5661.0


def test_to_float_empty_string():
    """Empty string should NOT coerce to 0 — Excel treats text-typed "" as
    #VALUE! in arithmetic, distinct from a truly empty cell (None -> 0).
    Conflating the two breaks IFERROR(""*x, fallback) handling."""
    from osheet.evaluator import _to_float
    import math
    assert math.isnan(_to_float(""))


def test_to_float_none_is_zero():
    """Empty CELL (None) should still coerce to 0 in arithmetic (unchanged)."""
    from osheet.evaluator import _to_float
    assert _to_float(None) == 0.0


def test_to_float_unparseable_returns_nan():
    from osheet.evaluator import _to_float
    import math
    assert math.isnan(_to_float("hello"))
    assert math.isnan(_to_float("N/A"))


def test_average_over_comma_formatted_strings():
    """Integration: AVERAGE works when cells contain '2,032' style strings."""
    import io, openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws['A1'] = '5,661'
    ws['B1'] = '2,032'
    ws['C1'] = '=AVERAGE(A1, B1)'
    buf = io.BytesIO(); wb.save(buf)

    from osheet.parser import parse_xlsx
    from osheet.analyzer import run_all
    from osheet.evaluator import evaluate_patch
    workbook = parse_xlsx(buf.getvalue())
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    c1 = next(c for c in workbook.all_cells if c.formula and "AVERAGE" in c.formula)
    assert abs(result[c1.stable_id] - (5661 + 2032) / 2) < 0.01


def test_sum_over_comma_formatted_range():
    """SUM over a range of comma-formatted strings."""
    import io, openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws['A1'] = '1,000'
    ws['A2'] = '(500)'  # accounting negative
    ws['A3'] = '2,500'
    ws['B1'] = '=SUM(A1:A3)'
    buf = io.BytesIO(); wb.save(buf)

    from osheet.parser import parse_xlsx
    from osheet.analyzer import run_all
    from osheet.evaluator import evaluate_patch
    workbook = parse_xlsx(buf.getvalue())
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    s = next(c for c in workbook.all_cells if c.formula and "SUM" in c.formula)
    assert abs(result[s.stable_id] - 3000) < 0.01


def test_offset_self_row_dependency_resolves_in_order():
    """OFFSET to a cell in the same row should force that cell to be evaluated first."""
    import io, openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Sheet1"
    ws['A1'] = 100
    # B1 depends on A1 via OFFSET, not direct ref
    ws['B1'] = "=OFFSET(B1,0,-1)+1"
    # C1 depends on B1 via OFFSET
    ws['C1'] = "=OFFSET(C1,0,-1)+1"
    # D1 depends on C1 via OFFSET
    ws['D1'] = "=OFFSET(D1,0,-1)+1"
    buf = io.BytesIO(); wb.save(buf)

    from osheet.parser import parse_xlsx
    from osheet.analyzer import run_all
    from osheet.evaluator import evaluate_patch
    workbook = parse_xlsx(buf.getvalue())
    run_all(workbook)
    result = evaluate_patch({}, workbook)

    b1 = next(c for c in workbook.all_cells if c.col == 2 and c.row == 1)
    c1 = next(c for c in workbook.all_cells if c.col == 3 and c.row == 1)
    d1 = next(c for c in workbook.all_cells if c.col == 4 and c.row == 1)
    # B1 = A1 + 1 = 101; C1 = B1 + 1 = 102; D1 = C1 + 1 = 103
    assert abs(result[b1.stable_id] - 101) < 0.01
    assert abs(result[c1.stable_id] - 102) < 0.01
    assert abs(result[d1.stable_id] - 103) < 0.01


def test_offset_with_static_constant_args_adds_dep():
    """Verify that _build_dep_graph captures OFFSET edges when args are constants."""
    import io, openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Sheet1"
    ws['A1'] = 1
    ws['B1'] = 2
    ws['C1'] = 3
    ws['D1'] = "=OFFSET(D1,0,-3)*10"  # references A1 via OFFSET
    buf = io.BytesIO(); wb.save(buf)

    from osheet.parser import parse_xlsx
    from osheet.analyzer import run_all
    from osheet.evaluator import evaluate_patch, _build_dep_graph
    workbook = parse_xlsx(buf.getvalue())
    run_all(workbook)

    deps = _build_dep_graph(workbook)
    d1 = next(c for c in workbook.all_cells if c.col == 4 and c.row == 1)
    a1 = next(c for c in workbook.all_cells if c.col == 1 and c.row == 1)
    assert a1.stable_id in deps[d1.stable_id]

    result = evaluate_patch({}, workbook)
    assert abs(result[d1.stable_id] - 10) < 0.01  # A1=1 * 10


def test_iferror_catches_empty_string_arithmetic():
    """=IFERROR(A1*0.15, " ") where A1=IFERROR(<err>, "") -> " ", not 0.0.

    This is the exact business_financial_plan pattern: a chained IFERROR where
    the inner branch returns "" and the outer branch multiplies it. Silent
    ""->0 coercion previously masked the arithmetic error, making IFERROR's
    fallback unreachable and returning 0.0 instead of " "."""
    import io, openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Sheet1"
    # A1 evaluates to "" via inner IFERROR; B1 then needs to surface that as
    # a #VALUE! error so its outer IFERROR can return the fallback.
    ws['A1'] = '=IFERROR(1/0, "")'
    ws['B1'] = '=IFERROR(A1*0.15, " ")'
    buf = io.BytesIO(); wb.save(buf)

    from osheet.parser import parse_xlsx
    from osheet.analyzer import run_all
    from osheet.evaluator import evaluate_patch
    workbook = parse_xlsx(buf.getvalue())
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    b1 = next(c for c in workbook.all_cells if c.row == 1 and c.col == 2)
    assert result[b1.stable_id] == " ", f"got {result[b1.stable_id]!r}"


def test_iferror_catches_text_arithmetic():
    """=IFERROR("hello"*2, "fallback") should return "fallback"."""
    import io, openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Sheet1"
    ws['A1'] = "hello"
    ws['B1'] = '=IFERROR(A1*2, "fallback")'
    buf = io.BytesIO(); wb.save(buf)

    from osheet.parser import parse_xlsx
    from osheet.analyzer import run_all
    from osheet.evaluator import evaluate_patch
    workbook = parse_xlsx(buf.getvalue())
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    b1 = next(c for c in workbook.all_cells if c.formula)
    assert result[b1.stable_id] == "fallback"


def test_empty_cell_arithmetic_still_zero():
    """=A1*5 where A1 is empty (None) should still return 0, not error.
    Empty CELL != empty STRING — only the latter should propagate as text."""
    import io, openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Sheet1"
    ws['B1'] = "=A1*5"
    buf = io.BytesIO(); wb.save(buf)

    from osheet.parser import parse_xlsx
    from osheet.analyzer import run_all
    from osheet.evaluator import evaluate_patch
    workbook = parse_xlsx(buf.getvalue())
    run_all(workbook)
    result = evaluate_patch({}, workbook)
    b1 = next(c for c in workbook.all_cells if c.formula)
    assert result[b1.stable_id] == 0
