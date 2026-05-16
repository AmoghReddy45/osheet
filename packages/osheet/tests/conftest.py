import io
import pytest
import openpyxl
from openpyxl.styles import PatternFill


@pytest.fixture
def simple_xlsx() -> bytes:
    """A minimal financial model xlsx for testing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revenue"

    # Header row
    ws["A1"] = "Month"
    ws["B1"] = "Revenue"
    ws["C1"] = "Costs"
    ws["D1"] = "Profit"

    # Assumption cell (yellow fill)
    ws["A3"] = "Churn Rate"
    ws["B3"] = 0.04
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["B3"].fill = yellow

    # Data rows
    for i, (rev, cost) in enumerate([(100, 60), (110, 65), (120, 70)], start=5):
        ws[f"A{i}"] = f"2026-0{i-4}"
        ws[f"B{i}"] = rev
        ws[f"C{i}"] = cost
        ws[f"D{i}"] = f"=B{i}-C{i}"

    # Output cell
    ws["B10"] = "=SUM(B5:B7)"
    ws["A10"] = "Total Revenue"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def cross_sheet_xlsx() -> bytes:
    """Workbook with cross-sheet formula references."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Inputs"
    ws1["A1"] = "Growth Rate"
    ws1["B1"] = 0.1

    ws2 = wb.create_sheet("Revenue")
    ws2["A1"] = "Base"
    ws2["B1"] = 1000
    ws2["A2"] = "Projected"
    ws2["B2"] = "=B1*(1+Inputs!B1)"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
