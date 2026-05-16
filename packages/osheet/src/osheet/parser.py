from __future__ import annotations
import io
from datetime import datetime as _dt
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell as OxlCell
from openpyxl.utils import column_index_from_string
from osheet.models import Cell, CellRole, NamedTable, Sheet, Workbook, Manifest


def _is_numeric_format(fmt: str | None) -> bool:
    """True if Excel cell number_format implies the value is numeric."""
    if not fmt or fmt == "General":
        return False
    # Common patterns: #,##0  #,##0.00  $#,##0  0%  0.00  #,##0_);(#,##0)  etc.
    # Heuristic: presence of #, 0, or % in the format, and no '@' (which means text)
    if "@" in fmt:
        return False
    return any(ch in fmt for ch in "#0%")


def _maybe_coerce_value(value, number_format: str | None, data_type: str | None = None):
    """If value is a string that parses as a number and the cell format is numeric,
    return the float; otherwise return value unchanged.

    If ``data_type`` is ``'s'`` (shared string) or ``'str'`` (inline string),
    Excel stored the value as TEXT — we leave it as-is so that aggregate
    functions (AVERAGE/SUM/etc) correctly skip these cells, matching Excel's
    text-skipping semantics.
    """
    if data_type in ("s", "str"):
        # Excel-stored text — preserve type even if the format is numeric.
        return value
    if not isinstance(value, str):
        return value
    if not _is_numeric_format(number_format):
        return value
    # Reuse the same coercion as evaluator
    s = value.strip()
    if not s:
        return value
    is_negative = False
    if s.startswith("(") and s.endswith(")"):
        is_negative = True
        s = s[1:-1].strip()
    for sym in ("$", "€", "£", "¥"):
        if s.startswith(sym):
            s = s[len(sym):].strip()
            break
    is_percent = s.endswith("%")
    if is_percent:
        s = s[:-1].strip()
    s = s.replace(",", "")
    try:
        v = float(s)
        if is_percent:
            v /= 100.0
        if is_negative:
            v = -v
        return v
    except (ValueError, TypeError):
        return value  # unparseable — keep as string


def _fill_color(cell: OxlCell) -> str | None:
    try:
        fill = cell.fill
        if fill and fill.fill_type not in (None, "none"):
            fg = fill.fgColor
            if fg and fg.type == "rgb" and fg.rgb not in ("00000000", "FFFFFFFF"):
                return fg.rgb
    except Exception:
        pass
    return None


def parse_xlsx(data: bytes) -> Workbook:
    wb_ox = load_workbook(io.BytesIO(data), data_only=False)
    sheets: list[Sheet] = []

    for ws in wb_ox.worksheets:
        cells: list[Cell] = []
        for row in ws.iter_rows():
            for ox_cell in row:
                if ox_cell.value is None:
                    continue
                raw_val = ox_cell.value
                formula: str | None = None
                value = raw_val
                if isinstance(raw_val, str) and raw_val.startswith("="):
                    formula = raw_val
                    value = None
                else:
                    # Coerce numeric-formatted text (e.g. "3,827", "(2,032)") to float
                    # so direct cell references propagate numbers, not strings.
                    # Excel-stored text-typed cells (data_type 's' or 'str') are
                    # preserved as strings so AVERAGE/SUM correctly skip them.
                    value = _maybe_coerce_value(
                        value, ox_cell.number_format, ox_cell.data_type
                    )

                cells.append(Cell(
                    stable_id=f"{ws.title}.{ox_cell.column_letter}{ox_cell.row}",
                    role=CellRole.UNKNOWN,
                    value=value,
                    formula=formula,
                    sheet_name=ws.title,
                    col=ox_cell.column,
                    row=ox_cell.row,
                    fill_color=_fill_color(ox_cell),
                ))

        sheets.append(Sheet(
            id=f"sheet.{ws.title.lower().replace(' ', '_')}",
            name=ws.title,
            cells=cells,
        ))

    # Collect Excel named tables (structured-reference targets) from each worksheet.
    named_tables: dict[str, NamedTable] = {}
    for ws in wb_ox.worksheets:
        for tname in ws.tables:
            tbl = ws.tables[tname]
            if isinstance(tbl, str):
                continue  # Skip legacy str-only entries (no column metadata)
            try:
                start_addr, end_addr = tbl.ref.split(":")
                start_col_letters = "".join(c for c in start_addr if c.isalpha())
                start_row = int("".join(c for c in start_addr if c.isdigit()))
                first_col = column_index_from_string(start_col_letters)
                end_col_letters = "".join(c for c in end_addr if c.isalpha())
                end_row = int("".join(c for c in end_addr if c.isdigit()))
                last_col = column_index_from_string(end_col_letters)
            except Exception:
                continue
            has_totals = bool(getattr(tbl, "totalsRowCount", 0) or 0)
            header_rows = getattr(tbl, "headerRowCount", 1) or 1
            first_data_row = start_row + header_rows
            last_data_row = end_row - (1 if has_totals else 0)
            cols: dict[str, int] = {}
            for i, col_def in enumerate(tbl.tableColumns or []):
                cols[col_def.name] = first_col + i
            named_tables[tname] = NamedTable(
                name=tname,
                sheet_name=ws.title,
                ref=tbl.ref,
                header_row=start_row,
                first_col=first_col,
                last_col=last_col,
                first_data_row=first_data_row,
                last_data_row=last_data_row,
                has_totals_row=has_totals,
                columns=cols,
            )

    manifest = Manifest(source_file="", sheet_count=len(sheets))
    # openpyxl exposes the workbook epoch as a datetime: 1899-12-30 for 1900
    # mode (default) or 1904-01-01 for 1904 mode (Mac Excel).
    try:
        manifest.epoch_year_1904 = (wb_ox.epoch == _dt(1904, 1, 1))
    except Exception:
        manifest.epoch_year_1904 = False
    return Workbook(sheets=sheets, manifest=manifest, named_tables=named_tables)
