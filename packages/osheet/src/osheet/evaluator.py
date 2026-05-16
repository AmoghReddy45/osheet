from __future__ import annotations

import math
import re
import sys
from datetime import date, datetime, time
from typing import Any

import numpy as np
import formulas
from openpyxl.utils import get_column_letter, column_index_from_string

from osheet.models import Workbook
from osheet.analyzer.graph import _parse_refs, _parse_range_endpoints


# ---------------------------------------------------------------------------
# Excel-faithful aggregate overrides.
#
# The `formulas` library's built-in AVERAGE/SUM/etc. don't match Excel when
# called with mixed text/numeric SCALAR arguments. E.g.
#   AVERAGE('3,827', 4908) returns 153512.0  (buggy comma parsing)
#   SUM('3,827', 4908)     returns #VALUE!
# whereas Excel skips text-typed scalars in aggregates and yields 4908 in
# both cases. We register Excel-faithful replacements at module import time
# via formulas.get_functions()[NAME] = fn.
#
# NOTE: SUMIF/SUMIFS/AVERAGEIF/AVERAGEIFS are intentionally NOT overridden;
# their range+criteria semantics are different and are already exercised by
# the existing test suite.
# ---------------------------------------------------------------------------


def _to_numeric_skip_text(args):
    """Flatten args; yield only numeric values, skip strings/None/text.
    Used by aggregates that match Excel's text-skipping semantics."""
    for a in args:
        if a is None:
            continue
        if isinstance(a, bool):
            yield float(a)  # TRUE=1, FALSE=0
        elif isinstance(a, (int, float, np.floating, np.integer)):
            fv = float(a)
            if not np.isnan(fv):
                yield fv
        elif isinstance(a, np.ndarray):
            for elem in a.flat:
                if isinstance(elem, bool) or isinstance(elem, np.bool_):
                    yield float(elem)
                elif isinstance(elem, (int, float, np.floating, np.integer)):
                    fv = float(elem)
                    if not np.isnan(fv):
                        yield fv
                # skip strings, None, NaN
        elif isinstance(a, str):
            continue  # Excel skips text-typed scalars in aggregates
        else:
            try:
                fv = float(a)
                if not np.isnan(fv):
                    yield fv
            except (ValueError, TypeError):
                continue


def _excel_average(*args):
    nums = list(_to_numeric_skip_text(args))
    if not nums:
        # Excel returns #DIV/0!. We return nan as a sentinel — the formulas
        # library and our _scalar()/result pipeline handle nan gracefully.
        return np.nan
    return sum(nums) / len(nums)


def _excel_sum(*args):
    return sum(_to_numeric_skip_text(args))


def _excel_min(*args):
    nums = list(_to_numeric_skip_text(args))
    return min(nums) if nums else 0


def _excel_max(*args):
    nums = list(_to_numeric_skip_text(args))
    return max(nums) if nums else 0


def _excel_count(*args):
    """COUNT: count of numeric values (ignores text, blanks, errors)."""
    return len(list(_to_numeric_skip_text(args)))


def _excel_counta(*args):
    """COUNTA: count of non-blank values including text."""
    n = 0
    for a in args:
        if a is None:
            continue
        if isinstance(a, np.ndarray):
            for elem in a.flat:
                if elem is None:
                    continue
                if isinstance(elem, str) and elem == "":
                    continue
                if isinstance(elem, float) and np.isnan(elem):
                    continue
                n += 1
        else:
            if isinstance(a, str) and a == "":
                continue
            if isinstance(a, float) and np.isnan(a):
                continue
            n += 1
    return n


def _excel_product(*args):
    nums = list(_to_numeric_skip_text(args))
    result = 1.0
    for v in nums:
        result *= v
    return result


# Register overrides at import time. These survive across multiple
# formulas.Parser() instances because get_functions() returns the
# library's global FUNCTIONS dict.
_excel_aggregates = {
    "AVERAGE": _excel_average,
    "SUM": _excel_sum,
    "MIN": _excel_min,
    "MAX": _excel_max,
    "COUNT": _excel_count,
    "COUNTA": _excel_counta,
    "PRODUCT": _excel_product,
}
for _name, _fn in _excel_aggregates.items():
    formulas.get_functions()[_name] = _fn


# Excel epoch (1900 date system). Excel's serial 1 is 1900-01-01, but because
# Excel incorrectly treats 1900 as a leap year, we use 1899-12-30 as the epoch
# so the resulting offsets match Excel for dates >= 1900-03-01 (i.e. all real
# financial-model usage).
_EXCEL_EPOCH = date(1899, 12, 30)
# 1904 mode is rare and none of the benchmark models use it. We track it on
# the Manifest but the conversion path below always uses _EXCEL_EPOCH. If a
# 1904-mode workbook surfaces in the future, thread the workbook epoch through
# to _to_float (TODO).


def _datetime_to_serial(d: Any, epoch: date = _EXCEL_EPOCH) -> float:
    """Convert datetime/date/time -> Excel serial number (days since epoch).
    Times become fractional days. Pre-1900-03-01 dates are off by 1 vs Excel
    (the 1900 leap-year bug); acceptable for financial models."""
    if isinstance(d, datetime):
        return (d.date() - epoch).days + (
            d.hour * 3600 + d.minute * 60 + d.second + d.microsecond / 1e6
        ) / 86400.0
    if isinstance(d, date):
        return (d - epoch).days
    if isinstance(d, time):
        return (
            d.hour * 3600 + d.minute * 60 + d.second + d.microsecond / 1e6
        ) / 86400.0
    raise TypeError(f"not a date/datetime/time: {type(d).__name__}")


# Single-cell ref (optionally sheet-qualified). Sheet may be quoted or bare.
_SINGLE_CELL_RE = re.compile(
    r"^\s*(?:(?:'([^']+)'|([A-Za-z_][\w ]*))!)?\s*\$?([A-Z]+)\$?(\d+)\s*$"
)


def _fkey(sheet_name: str, row: int, col: int) -> str:
    """Normalized cell key: SHEETNAME!COLROW (uppercase, no $)."""
    return f"{sheet_name.upper()}!{get_column_letter(col)}{row}"


def _normalize_input_key(input_key: str) -> str:
    """Normalize formulas-library input key to match our value_map keys.

    formulas lib outputs: 'FIXED ASSETS'!D15 (with quotes for sheet names with spaces)
    Our _fkey() outputs:  FIXED ASSETS!D15   (no quotes)

    Also handles the range-emitted form ``(SHEET!A1: SHEET!B2)`` with wrapping
    parens and whitespace around the colon.
    """
    s = input_key.strip()
    # Strip wrapping parens (formulas lib emits e.g. "(TBA!D10: TBA!I10)")
    if s.startswith("(") and s.endswith(")"):
        s = s[1:-1].strip()
    # Remove whitespace around ":" so downstream string-splitting works.
    s = re.sub(r"\s*:\s*", ":", s)
    if s.startswith("'"):
        bang = s.find("'", 1)
        if bang != -1 and s[bang + 1 : bang + 2] == "!":
            sheet = s[1:bang]
            rest = s[bang + 2 :]
            return f"{sheet}!{rest}"
    return s


def _expand_range(range_key: str, default_sheet: str, value_map: dict[str, Any]) -> np.ndarray:
    """Turn 'B2:B13' or 'SHEET!B2:SHEET!B13' into a numpy row array for SUM etc.

    Tolerant of all forms emitted by the formulas library and authored Excel
    expressions: bare ranges, mixed quoting (``TBA!$D$10:'TBA'!I10``), repeated
    sheet prefixes on both sides, quoted sheet names with spaces, and wrapping
    parens with whitespace around the colon.

    Preserves string values (uses an object dtype) so that text-criteria
    aggregations such as SUMIF can match labels in the criteria range."""
    range_key = _normalize_input_key(range_key)
    endpoints = _parse_range_endpoints(range_key, default_sheet)
    if endpoints is None:
        return np.array([[np.nan]])
    sheet, c1, r1, c2, r2 = endpoints

    vals: list[Any] = []
    has_unparseable_string = False
    for row in range(min(r1, r2), max(r1, r2) + 1):
        for col in range(min(c1, c2), max(c1, c2) + 1):
            v = value_map.get(f"{sheet.upper()}!{get_column_letter(col)}{row}")
            if isinstance(v, str):
                # Attempt Excel-style numeric coercion (e.g. "5,661" -> 5661,
                # "(2,032)" -> -2032). If the string is genuinely textual
                # (e.g. SUMIF criterion like "X", a label like "Total"),
                # keep it as-is so criteria matching still works.
                coerced = _coerce_string_to_float(v)
                if math.isnan(coerced):
                    vals.append(v)
                    has_unparseable_string = True
                else:
                    vals.append(coerced)
            elif v is None:
                vals.append(0.0)  # Excel treats blanks as 0 in numeric contexts
            else:
                vals.append(_to_float(v))
    if has_unparseable_string:
        return np.array([vals], dtype=object)
    return np.array([vals])


def _coerce_string_to_float(s: str) -> float:
    """Attempt Excel-compatible numeric coercion. Returns nan if not parseable.

    Handles common Excel-formatted text forms that Excel implicitly coerces
    to numbers during arithmetic:
      - Thousands separators:        "5,661" -> 5661.0
      - Accounting-format negatives: "(2,032)" -> -2032.0
      - Currency prefixes:           "$5,661" -> 5661.0
      - Percentages:                 "50%" -> 0.5
      - Leading/trailing whitespace: "  5,661  " -> 5661.0
      - Scientific notation:         "5.66e3" -> 5660.0 (via float())

    Note: empty string ("") returns nan (NOT 0.0). Excel distinguishes empty
    cells (treated as 0 in arithmetic) from text values including "" (which
    raise #VALUE! in arithmetic and can be caught by IFERROR). Conflating the
    two breaks the common =IFERROR(<inner-that-may-return-"">*x, fallback)
    pattern: silent 0-coercion makes IFERROR's error path unreachable.
    """
    s = s.strip()
    if not s:
        return float("nan")  # empty string -> not coerceable; let arithmetic raise #VALUE!
    is_negative = False
    # Accounting format: (1,234) means -1234
    if s.startswith("(") and s.endswith(")"):
        is_negative = True
        s = s[1:-1].strip()
    # Strip currency symbols (extend as needed)
    for sym in ("$", "€", "£", "¥"):
        if s.startswith(sym):
            s = s[len(sym):].strip()
            break
    # Percentage
    is_percent = s.endswith("%")
    if is_percent:
        s = s[:-1].strip()
    # Remove thousands separators
    s = s.replace(",", "")
    try:
        v = float(s)
        if is_percent:
            v /= 100.0
        if is_negative:
            v = -v
        return v
    except (ValueError, TypeError):
        return float("nan")


def _to_float(val: Any) -> float:
    if val is None:
        return 0.0  # Excel treats empty/missing cells as 0 in arithmetic
    if isinstance(val, (datetime, date, time)):
        # Dates pass through formula evaluation as Excel serial numbers so that
        # YEAR/EDATE/EOMONTH/TEXT/etc. (which the formulas lib implements
        # against serials) work, instead of erroring with #VALUE!.
        return _datetime_to_serial(val)
    if isinstance(val, bool):
        return float(val)
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        return _coerce_string_to_float(val)
    # Try numpy scalar extraction
    try:
        arr = np.array(val)
        return float(arr.item())
    except Exception:
        pass
    if hasattr(val, "__iter__") and not isinstance(val, (str, bytes)):
        try:
            return float(list(list(val)[0])[0])
        except Exception:
            return np.nan
    try:
        return float(val)
    except (TypeError, ValueError):
        return np.nan


def _scalar(val: Any) -> Any:
    """Extract a Python scalar from a formulas Array or numpy array."""
    if val is None:
        return None
    if isinstance(val, (int, float, bool)):
        return val
    # Try numpy scalar extraction first (works for 0-d arrays and Array objects)
    try:
        arr = np.array(val)
        return arr.item()
    except Exception:
        pass
    # Fallback: iterate
    if hasattr(val, "__iter__") and not isinstance(val, (str, bytes)):
        try:
            inner = list(val)
            if inner and hasattr(inner[0], "__iter__"):
                inner = list(inner[0])
            v = inner[0]
            return float(v) if isinstance(v, (int, float, np.floating)) else v
        except (IndexError, TypeError):
            return None
    return val


def _split_top_level_commas(body: str) -> list[str]:
    """Split a function body at top-level commas, respecting parens and quotes."""
    parts: list[str] = []
    depth = 0
    in_quote = False
    start = 0
    i = 0
    while i < len(body):
        ch = body[i]
        if ch == '"':
            in_quote = not in_quote
        elif not in_quote:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
            elif ch == "," and depth == 0:
                parts.append(body[start:i])
                start = i + 1
        i += 1
    parts.append(body[start:])
    return parts


def _find_matching_paren(text: str, open_idx: int) -> int:
    """Given index of '(' in text, return index of matching ')'. -1 if not found.
    Ignores parens inside double-quoted string literals."""
    depth = 0
    in_quote = False
    i = open_idx
    while i < len(text):
        ch = text[i]
        if ch == '"':
            in_quote = not in_quote
        elif not in_quote:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
                if depth == 0:
                    return i
        i += 1
    return -1


def _find_innermost_offset(formula: str) -> tuple[int, int] | None:
    """Find an OFFSET(...) call whose body contains no nested OFFSET(.
    Returns (start_of_OFFSET, index_of_closing_paren) or None."""
    upper = formula.upper()
    search_from = 0
    while True:
        idx = upper.find("OFFSET(", search_from)
        if idx == -1:
            return None
        # Must be at start of token (prev char non-alnum/underscore) to avoid matching e.g. MYOFFSET(
        if idx > 0 and (formula[idx - 1].isalnum() or formula[idx - 1] == "_"):
            search_from = idx + 1
            continue
        open_paren = idx + len("OFFSET")  # points at '('
        close_paren = _find_matching_paren(formula, open_paren)
        if close_paren == -1:
            return None
        body = formula[open_paren + 1 : close_paren]
        if "OFFSET(" not in body.upper():
            return (idx, close_paren)
        # Has nested OFFSET; recurse into body
        inner = _find_innermost_offset(body)
        if inner is None:
            # Shouldn't happen, but be safe
            return (idx, close_paren)
        b_start, b_end = inner
        return (open_paren + 1 + b_start, open_paren + 1 + b_end)


def _eval_subexpr(expr: str, default_sheet: str, value_map: dict[str, Any], parser: Any) -> int:
    expr = _collapse_redundant_sheet_in_range(expr)
    func = parser.ast(f"={expr}")[1].compile()
    kwargs: dict[str, Any] = {}
    sheet_prefix = default_sheet.upper() + "!"
    for input_key in func.inputs:
        normalized = _normalize_input_key(input_key)
        if ":" in normalized:
            kwargs[input_key] = _expand_range(normalized, default_sheet.upper(), value_map)
        else:
            lookup = normalized if "!" in normalized else sheet_prefix + normalized
            v = value_map.get(lookup)
            if isinstance(v, (int, float, type(None), datetime, date, time)):
                kwargs[input_key] = _to_float(v)
            elif isinstance(v, str):
                coerced = _coerce_string_to_float(v)
                kwargs[input_key] = v if math.isnan(coerced) else coerced
            else:
                kwargs[input_key] = v
    result = _scalar(func(**kwargs))
    return int(result)


def _eval_subexpr_scalar(
    expr: str,
    default_sheet: str,
    value_map: dict[str, Any],
    parser: Any,
    named_tables: dict[str, Any] | None,
    cell_row: int,
) -> Any:
    """Evaluate a sub-expression, returning the raw scalar.

    Unlike ``_eval_subexpr`` this does not cast to int and additionally
    resolves structured table references and OFFSET calls in ``expr``
    before parsing. Any failure returns ``None``.
    """
    try:
        text = expr
        if named_tables and "[" in text:
            text = _resolve_structured_refs(text, cell_row, named_tables)
        if "SUBTOTAL(" in text.upper():
            text = _rewrite_subtotal(text)
        if "OFFSET" in text.upper():
            text = _resolve_offset_in_formula(text, default_sheet, value_map, parser)
        text = _collapse_redundant_sheet_in_range(text)
        if not text.startswith("="):
            text = "=" + text
        func = parser.ast(text)[1].compile()
        kwargs: dict[str, Any] = {}
        sheet_prefix = default_sheet.upper() + "!"
        for input_key in func.inputs:
            normalized = _normalize_input_key(input_key)
            if ":" in normalized:
                kwargs[input_key] = _expand_range(normalized, default_sheet.upper(), value_map)
            else:
                lookup = normalized if "!" in normalized else sheet_prefix + normalized
                v = value_map.get(lookup)
                if isinstance(v, (int, float, type(None), datetime, date, time)):
                    kwargs[input_key] = _to_float(v)
                elif isinstance(v, str):
                    coerced = _coerce_string_to_float(v)
                    kwargs[input_key] = v if math.isnan(coerced) else coerced
                else:
                    kwargs[input_key] = v
        return _scalar(func(**kwargs))
    except Exception:
        return None


def _try_if_short_circuit(
    formula: str,
    default_sheet: str,
    value_map: dict[str, Any],
    parser: Any,
    named_tables: dict[str, Any] | None,
    cell_row: int,
) -> tuple[bool, Any]:
    """Short-circuit IF(cond, true, false) when an unresolvable OFFSET sits
    in the inactive branch.

    Returns ``(handled, value)``. ``handled=True`` means the result is
    authoritative and the caller should return ``value`` directly.
    ``handled=False`` means the caller should fall back to normal evaluation.
    """
    body = formula.lstrip()
    if body.startswith("="):
        body = body[1:].lstrip()
    if not body.upper().startswith("IF("):
        return (False, None)
    open_paren = body.upper().find("IF(") + 2  # index of '('
    close = _find_matching_paren(body, open_paren)
    if close == -1 or close != len(body.rstrip()) - 1:
        return (False, None)
    inner = body[open_paren + 1 : close]
    parts = _split_top_level_commas(inner)
    if len(parts) != 3:
        return (False, None)
    cond_expr, true_expr, false_expr = (p.strip() for p in parts)
    cond_val = _eval_subexpr_scalar(
        cond_expr, default_sheet, value_map, parser, named_tables, cell_row
    )
    if cond_val is None:
        return (False, None)
    if cond_val == 0 or cond_val is False or cond_val == "":
        is_true = False
    else:
        is_true = bool(cond_val)
    active = true_expr if is_true else false_expr
    active = active.strip()
    # Recurse into nested IF
    if active.upper().startswith("IF("):
        recur_close = _find_matching_paren(active, 2)  # '(' is at index 2
        if recur_close == len(active) - 1:
            handled, val = _try_if_short_circuit(
                "=" + active, default_sheet, value_map, parser, named_tables, cell_row
            )
            if handled:
                return (True, val)
    return (
        True,
        _eval_subexpr_scalar(
            active, default_sheet, value_map, parser, named_tables, cell_row
        ),
    )


def _parse_single_cell_ref(ref: str) -> tuple[str | None, str, int]:
    """Parse 'A1' / '$A$1' / 'Sheet!A1' / ''Sheet Name'!A1' into (sheet_or_None, col_letters, row).
    Raises ValueError if not a single-cell reference."""
    m = _SINGLE_CELL_RE.match(ref)
    if not m:
        raise ValueError(f"not a single cell ref: {ref!r}")
    quoted_sheet, bare_sheet, col_letters, row_str = m.groups()
    sheet = quoted_sheet or bare_sheet
    return sheet, col_letters.upper(), int(row_str)


# Matches `<sheet>!<cell>:<sheet>!<cell>` where sheets may be quoted or bare on
# either side independently. Anchored at non-identifier boundaries so we don't
# eat surrounding tokens (e.g. ``=SUM(``).
_REDUNDANT_SHEET_RE = re.compile(
    r"(?<![A-Za-z0-9_$])"
    r"(?:'([^']+)'|([A-Za-z_][\w ]*))!(\$?[A-Z]+\$?\d+)"
    r"\s*:\s*"
    r"(?:'([^']+)'|([A-Za-z_][\w ]*))!(\$?[A-Z]+\$?\d+)"
    r"(?![A-Za-z0-9_$])"
)


def _collapse_redundant_sheet_in_range(text: str) -> str:
    """Collapse ``Sheet!X:Sheet!Y`` (any quoting combo) → ``Sheet!X:Y``.

    The left-side prefix wins so quoting style on the left is preserved
    verbatim. This is required because the ``formulas`` library treats the
    duplicated-sheet form as two scalars rather than a range.
    """
    def repl(m: re.Match) -> str:
        left_sheet = m.group(1) or m.group(2)
        right_sheet = m.group(4) or m.group(5)
        if left_sheet.upper() == right_sheet.upper():
            # Use the original left-prefix (preserves quoting and `$` anchors).
            left_prefix = m.group(0).split(":", 1)[0].rstrip()
            return f"{left_prefix}:{m.group(6)}"
        return m.group(0)
    return _REDUNDANT_SHEET_RE.sub(repl, text)


def _resolve_offset_in_formula(
    formula: str,
    default_sheet: str,
    value_map: dict[str, Any],
    parser: Any,
) -> str:
    """Rewrite OFFSET(ref, rows, cols, [height], [width]) → concrete cell or range
    in the formula text. On any failure, returns the original formula unchanged."""
    try:
        current = formula
        # Bound iterations to avoid pathological loops
        for _ in range(64):
            upper = current.upper()
            if "OFFSET(" not in upper:
                break
            located = _find_innermost_offset(current)
            if located is None:
                break
            offset_start, close_paren = located
            open_paren = offset_start + len("OFFSET")
            body = current[open_paren + 1 : close_paren]
            args = _split_top_level_commas(body)
            if len(args) < 3:
                return formula  # malformed; bail
            ref_arg = args[0].strip()
            rows_int = _eval_subexpr(args[1].strip(), default_sheet, value_map, parser)
            cols_int = _eval_subexpr(args[2].strip(), default_sheet, value_map, parser)
            height = (
                _eval_subexpr(args[3].strip(), default_sheet, value_map, parser)
                if len(args) >= 4 and args[3].strip()
                else 1
            )
            width = (
                _eval_subexpr(args[4].strip(), default_sheet, value_map, parser)
                if len(args) >= 5 and args[4].strip()
                else 1
            )

            sheet_name, col_letters, base_row = _parse_single_cell_ref(ref_arg)
            base_col = column_index_from_string(col_letters)
            new_col = base_col + cols_int
            new_row = base_row + rows_int
            if new_col < 1 or new_row < 1:
                raise ValueError("OFFSET out of range")
            top_addr = f"{get_column_letter(new_col)}{new_row}"
            if height > 1 or width > 1:
                bottom_col = new_col + max(width, 1) - 1
                bottom_row = new_row + max(height, 1) - 1
                bottom_addr = f"{get_column_letter(bottom_col)}{bottom_row}"
                new_addr = f"{top_addr}:{bottom_addr}"
            else:
                new_addr = top_addr
            if sheet_name:
                prefix = (
                    f"'{sheet_name}'!" if " " in sheet_name else f"{sheet_name}!"
                )
                new_addr = prefix + new_addr

            current = current[:offset_start] + new_addr + current[close_paren + 1 :]

        # Normalize `Sheet!X:Sheet!Y` → `Sheet!X:Y` (formulas lib treats the
        # duplicated-sheet form as two scalars, not a range).
        return _collapse_redundant_sheet_in_range(current)
    except Exception:
        return formula


# Structured-reference pattern. Order of alternatives matters - longer/more
# specific patterns come first so the regex engine prefers them.
_STRUCTURED_REF_RE = re.compile(
    r"\b([A-Za-z_]\w*)"
    r"(?:"
    r"\[\[\#This\sRow\],\[([^\]]+)\]:\[([^\]]+)\]\]"   # groups 2,3: #This Row col range
    r"|\[\[\#This\sRow\],\[([^\]]+)\]\]"               # group 4: #This Row, single col
    r"|\[\[\#Totals\],\[([^\]]+)\]:\[([^\]]+)\]\]"     # groups 5,6: #Totals col range
    r"|\[\[\#Totals\],\[([^\]]+)\]\]"                  # group 7: #Totals single col
    r"|\[\[\#Headers\],\[([^\]]+)\]\]"                 # group 8: #Headers single col
    r"|\[\[([^\]\#][^\]]*)\]:\[([^\]]+)\]\]"           # groups 9,10: col range [[A]:[B]]
    r"|\[\#All\]"
    r"|\[\#Data\]"
    r"|\[\#Headers\]"
    r"|\[\#Totals\]"
    r"|\[([^\]\#\[][^\]]*)\]"                          # group 11: bare [Col]
    r")"
)


def _sheet_prefix(sheet_name: str) -> str:
    return f"'{sheet_name}'!" if " " in sheet_name or "'" in sheet_name else f"{sheet_name}!"


# SUBTOTAL function-code -> plain aggregate function name. The 1xx codes ignore
# hidden rows; without UI state we treat both code ranges identically.
_SUBTOTAL_FN_BY_CODE = {
    1: "AVERAGE", 101: "AVERAGE",
    2: "COUNT", 102: "COUNT",
    3: "COUNTA", 103: "COUNTA",
    4: "MAX", 104: "MAX",
    5: "MIN", 105: "MIN",
    6: "PRODUCT", 106: "PRODUCT",
    7: "STDEV", 107: "STDEV",
    8: "STDEVP", 108: "STDEVP",
    9: "SUM", 109: "SUM",
    10: "VAR", 110: "VAR",
    11: "VARP", 111: "VARP",
}


def _rewrite_subtotal(formula: str) -> str:
    """Rewrite SUBTOTAL(code, args...) -> <AGGREGATE>(args...) so that the
    formulas library (which does not implement SUBTOTAL) can evaluate it.

    Iterates innermost-first so nested SUBTOTALs are handled correctly."""
    upper = formula.upper()
    if "SUBTOTAL(" not in upper:
        return formula
    # Repeatedly find innermost SUBTOTAL and rewrite it.
    current = formula
    for _ in range(64):
        upper = current.upper()
        idx = -1
        # Find an occurrence of SUBTOTAL( whose body contains no inner SUBTOTAL(
        search_from = 0
        while True:
            pos = upper.find("SUBTOTAL(", search_from)
            if pos == -1:
                break
            if pos > 0 and (current[pos - 1].isalnum() or current[pos - 1] == "_"):
                search_from = pos + 1
                continue
            open_paren = pos + len("SUBTOTAL")
            close = _find_matching_paren(current, open_paren)
            if close == -1:
                return current
            body = current[open_paren + 1: close]
            if "SUBTOTAL(" in body.upper():
                search_from = pos + 1
                continue
            idx = pos
            break
        if idx == -1:
            return current
        open_paren = idx + len("SUBTOTAL")
        close = _find_matching_paren(current, open_paren)
        body = current[open_paren + 1: close]
        args = _split_top_level_commas(body)
        if not args:
            return current
        try:
            code = int(float(args[0].strip()))
        except (ValueError, TypeError):
            return current
        fn = _SUBTOTAL_FN_BY_CODE.get(code)
        if fn is None:
            return current
        rest = ",".join(args[1:])
        current = current[:idx] + fn + "(" + rest + ")" + current[close + 1:]
    return current


def _resolve_structured_refs(
    formula: str,
    cell_row: int,
    named_tables: dict[str, Any],
) -> str:
    """Rewrite structured Table references to A1-style ranges.

    Supports:
      - Table[[#This Row],[Col]]
      - Table[Col]                          (whole data column)
      - Table[[Col1]:[Col2]]                (multi-column data area)
      - Table[[#Totals],[Col]]              (totals row cell)
      - Table[[#Totals],[Col1]:[Col2]]      (totals row range)
      - Table[[#Headers],[Col]]             (header cell)
      - Table[#All|#Data|#Headers|#Totals]  (whole-table keywords)

    Returns the formula unchanged if no rewrite is possible (so the formula
    evaluator will surface None for the affected cells)."""
    if not named_tables or "[" not in formula:
        return formula

    def _sub(m: re.Match) -> str:
        tname = m.group(1)
        tbl = named_tables.get(tname)
        if tbl is None:
            return m.group(0)
        prefix = _sheet_prefix(tbl.sheet_name)
        full_match = m.group(0)

        # #This Row, col range
        if m.group(2) and m.group(3):
            c1, c2 = m.group(2), m.group(3)
            if c1 not in tbl.columns or c2 not in tbl.columns:
                return full_match
            return (
                f"{prefix}{get_column_letter(tbl.columns[c1])}{cell_row}"
                f":{get_column_letter(tbl.columns[c2])}{cell_row}"
            )

        # #This Row, single col
        if m.group(4):
            col_name = m.group(4)
            if col_name not in tbl.columns:
                return full_match
            return f"{prefix}{get_column_letter(tbl.columns[col_name])}{cell_row}"

        # #Totals, col range
        if m.group(5) and m.group(6):
            if not tbl.has_totals_row:
                return full_match
            c1, c2 = m.group(5), m.group(6)
            if c1 not in tbl.columns or c2 not in tbl.columns:
                return full_match
            totals_row = tbl.last_data_row + 1
            return (
                f"{prefix}{get_column_letter(tbl.columns[c1])}{totals_row}"
                f":{get_column_letter(tbl.columns[c2])}{totals_row}"
            )

        # #Totals, single col
        if m.group(7):
            if not tbl.has_totals_row:
                return full_match
            col_name = m.group(7)
            if col_name not in tbl.columns:
                return full_match
            totals_row = tbl.last_data_row + 1
            return f"{prefix}{get_column_letter(tbl.columns[col_name])}{totals_row}"

        # #Headers, single col
        if m.group(8):
            col_name = m.group(8)
            if col_name not in tbl.columns:
                return full_match
            return f"{prefix}{get_column_letter(tbl.columns[col_name])}{tbl.header_row}"

        # Column range [[A]:[B]]
        if m.group(9) and m.group(10):
            c1, c2 = m.group(9), m.group(10)
            if c1 not in tbl.columns or c2 not in tbl.columns:
                return full_match
            return (
                f"{prefix}{get_column_letter(tbl.columns[c1])}{tbl.first_data_row}"
                f":{get_column_letter(tbl.columns[c2])}{tbl.last_data_row}"
            )

        # [#All] / [#Data] / [#Headers] / [#Totals] - whole-table keywords
        rest = full_match[len(tname):]
        if rest == "[#All]":
            last_row = tbl.last_data_row + (1 if tbl.has_totals_row else 0)
            return (
                f"{prefix}{get_column_letter(tbl.first_col)}{tbl.header_row}"
                f":{get_column_letter(tbl.last_col)}{last_row}"
            )
        if rest == "[#Data]":
            return (
                f"{prefix}{get_column_letter(tbl.first_col)}{tbl.first_data_row}"
                f":{get_column_letter(tbl.last_col)}{tbl.last_data_row}"
            )
        if rest == "[#Headers]":
            return (
                f"{prefix}{get_column_letter(tbl.first_col)}{tbl.header_row}"
                f":{get_column_letter(tbl.last_col)}{tbl.header_row}"
            )
        if rest == "[#Totals]":
            if not tbl.has_totals_row:
                return full_match
            totals_row = tbl.last_data_row + 1
            return (
                f"{prefix}{get_column_letter(tbl.first_col)}{totals_row}"
                f":{get_column_letter(tbl.last_col)}{totals_row}"
            )

        # Bare [ColName] - whole data column range
        if m.group(11):
            col_name = m.group(11)
            if col_name not in tbl.columns:
                return full_match
            col = tbl.columns[col_name]
            if tbl.last_data_row < tbl.first_data_row:
                # Empty table - emit single cell to avoid an inverted range
                return f"{prefix}{get_column_letter(col)}{tbl.first_data_row}"
            return (
                f"{prefix}{get_column_letter(col)}{tbl.first_data_row}"
                f":{get_column_letter(col)}{tbl.last_data_row}"
            )

        return full_match

    return _STRUCTURED_REF_RE.sub(_sub, formula)


def _extract_offset_deps(
    formula: str,
    default_sheet: str,
    static_value_map: dict[str, Any],
    parser: Any,
) -> list[str]:
    """Find OFFSET expressions in a formula and return the value_map-style
    refs they would point to, based on a static value_map of constants.

    Returns refs as ``'SHEET!COLROW'`` (matching ``_fkey`` key format). For
    OFFSETs whose row/col args can't be evaluated statically (e.g. they
    reference other formula cells), the corresponding entry is omitted —
    no worse than the existing behavior.
    """
    deps: list[str] = []
    if "OFFSET" not in formula.upper():
        return deps
    upper = formula.upper()
    i = 0
    while True:
        idx = upper.find("OFFSET(", i)
        if idx == -1:
            break
        # Must be at start of token (prev char non-alnum/underscore) to avoid
        # matching e.g. MYOFFSET(.
        if idx > 0 and (formula[idx - 1].isalnum() or formula[idx - 1] == "_"):
            i = idx + 1
            continue
        open_paren = idx + len("OFFSET")
        close = _find_matching_paren(formula, open_paren)
        if close == -1:
            break
        body = formula[open_paren + 1 : close]
        parts = _split_top_level_commas(body)
        if len(parts) < 3:
            i = close + 1
            continue
        ref_text = parts[0].strip()
        try:
            sheet, col_letters, base_row = _parse_single_cell_ref(ref_text)
            base_col = column_index_from_string(col_letters)
            rows_val = _eval_subexpr(
                parts[1].strip(), default_sheet, static_value_map, parser
            )
            cols_val = _eval_subexpr(
                parts[2].strip(), default_sheet, static_value_map, parser
            )
            new_col = base_col + int(cols_val)
            new_row = base_row + int(rows_val)
            if new_col >= 1 and new_row >= 1:
                target_sheet = (sheet or default_sheet).upper()
                deps.append(
                    f"{target_sheet}!{get_column_letter(new_col)}{new_row}"
                )
        except Exception:
            pass
        i = close + 1
    return deps


def _build_dep_graph(
    workbook: Workbook,
) -> dict[str, set[str]]:
    """
    Build {stable_id -> set of stable_ids it depends on} using formula text,
    bypassing the (potentially stale) cell.depends_on field.

    Also includes OFFSET-discovered edges when the OFFSET row/col args are
    statically resolvable (i.e. reference non-formula constant cells). This
    ensures Tarjan orders OFFSET targets before their consumers — without
    this, an OFFSET target may be evaluated as 0 because its source hasn't
    been computed yet.
    """
    parser = formulas.Parser()

    # Position lookup: (sheet_name, col, row) -> stable_id
    pos_to_id: dict[tuple[str, int, int], str] = {}
    # Key lookup: SHEET!COLROW -> stable_id (matches _fkey output)
    key_to_id: dict[str, str] = {}
    # Static value_map seeded only with non-formula (constant) cells. OFFSET
    # arg evaluation against this map will fail for cells whose args
    # reference formula cells — those OFFSETs are then skipped here, matching
    # the original (pre-fix) dep-graph behavior.
    static_value_map: dict[str, Any] = {}
    for sheet in workbook.sheets:
        for cell in sheet.cells:
            pos_to_id[(sheet.name, cell.col, cell.row)] = cell.stable_id
            k = _fkey(sheet.name, cell.row, cell.col)
            key_to_id[k] = cell.stable_id
            if cell.formula is None:
                static_value_map[k] = cell.value

    deps: dict[str, set[str]] = {}
    for sheet in workbook.sheets:
        for cell in sheet.cells:
            if not cell.formula:
                continue
            refs = _parse_refs(cell.formula, sheet.name)
            dep_ids: set[str] = set()
            for (sname, col, row) in refs:
                sid = pos_to_id.get((sname, col, row))
                if sid:
                    dep_ids.add(sid)

            # Include OFFSET targets when row/col args are statically
            # resolvable (constants from non-formula cells).
            offset_refs = _extract_offset_deps(
                cell.formula,
                sheet.name.upper(),
                static_value_map,
                parser,
            )
            for ref_key in offset_refs:
                sid = key_to_id.get(ref_key)
                if sid and sid != cell.stable_id:  # don't create self-edges
                    dep_ids.add(sid)

            deps[cell.stable_id] = dep_ids
    return deps


def _tarjan_sccs(deps: dict[str, set[str]]) -> list[list[str]]:
    """Return list of SCCs using Tarjan's algorithm.

    Iteration order of ``all_nodes`` (a set) is non-deterministic in general,
    but because Tarjan's DFS recurses into dependencies before finishing a node,
    dependency SCCs are emitted (appended) before the SCCs that depend on them.
    The result is therefore in sources-first topological order for this
    implementation — suitable for direct use in evaluate_patch without reversal.

    SCCs with len > 1, or len==1 with a self-loop, are circular.
    """
    all_nodes: set[str] = set(deps.keys())
    for s in deps.values():
        all_nodes |= s

    index_counter = [0]
    stack: list[str] = []
    lowlink: dict[str, int] = {}
    index: dict[str, int] = {}
    on_stack: dict[str, bool] = {}
    sccs: list[list[str]] = []

    def strongconnect(v: str) -> None:
        index[v] = lowlink[v] = index_counter[0]
        index_counter[0] += 1
        stack.append(v)
        on_stack[v] = True
        for w in deps.get(v, set()):
            if w not in index:
                strongconnect(w)
                lowlink[v] = min(lowlink[v], lowlink[w])
            elif on_stack.get(w, False):
                lowlink[v] = min(lowlink[v], index[w])
        if lowlink[v] == index[v]:
            scc: list[str] = []
            while True:
                w = stack.pop()
                on_stack[w] = False
                scc.append(w)
                if w == v:
                    break
            sccs.append(scc)

    old_limit = sys.getrecursionlimit()
    sys.setrecursionlimit(max(old_limit, len(all_nodes) * 3 + 1000))
    try:
        for v in all_nodes:
            if v not in index:
                strongconnect(v)
    finally:
        sys.setrecursionlimit(old_limit)
    return sccs


def _eval_one_cell(
    sheet_name: str,
    cell: Any,
    value_map: dict[str, Any],
    id_to_key: dict[str, str],
    parser: Any,
    named_tables: dict[str, Any] | None = None,
) -> Any:
    """Evaluate a single formula cell. Returns scalar or None on error."""
    try:
        formula_text = cell.formula
        if named_tables and "[" in formula_text:
            formula_text = _resolve_structured_refs(formula_text, cell.row, named_tables)
        if "SUBTOTAL(" in formula_text.upper():
            formula_text = _rewrite_subtotal(formula_text)
        if "OFFSET" in formula_text.upper():
            formula_text = _resolve_offset_in_formula(
                formula_text, sheet_name.upper(), value_map, parser
            )
            # If the rewrite couldn't resolve every OFFSET (e.g. the shifted
            # column/row would fall out of range) AND the outer wrapper is an
            # IF, short-circuit the IF and skip the dead branch. Excel never
            # evaluates that branch, and the unresolved OFFSET would otherwise
            # cause parse/dispatch failure.
            if "OFFSET(" in formula_text.upper():
                stripped = formula_text.lstrip("=").lstrip()
                if stripped.upper().startswith("IF("):
                    handled, val = _try_if_short_circuit(
                        formula_text,
                        sheet_name.upper(),
                        value_map,
                        parser,
                        named_tables,
                        cell.row,
                    )
                    if handled:
                        return val
        formula_text = _collapse_redundant_sheet_in_range(formula_text)
        func = parser.ast(formula_text)[1].compile()
        kwargs: dict[str, Any] = {}
        sheet_prefix = sheet_name.upper() + "!"
        for input_key in func.inputs:
            normalized = _normalize_input_key(input_key)
            if ":" in normalized:
                kwargs[input_key] = _expand_range(normalized, sheet_name.upper(), value_map)
            else:
                lookup_key = normalized if "!" in normalized else sheet_prefix + normalized
                v = value_map.get(lookup_key)
                if isinstance(v, (int, float, type(None), datetime, date, time)):
                    kwargs[input_key] = _to_float(v)
                elif isinstance(v, str):
                    # Excel-style implicit numeric coercion of formatted text
                    # (e.g. "5,661" or "(2,032)"). If unparseable, keep the
                    # raw string so criterion-style usage (e.g. SUMIF) works.
                    coerced = _coerce_string_to_float(v)
                    kwargs[input_key] = v if math.isnan(coerced) else coerced
                else:
                    kwargs[input_key] = v
        return _scalar(func(**kwargs))
    except Exception:
        return None


def _gauss_seidel(
    scc_cells: list[tuple[str, Any]],
    value_map: dict[str, Any],
    id_to_key: dict[str, str],
    parser: Any,
    max_iterations: int = 50,
    tolerance: float = 1e-6,
    named_tables: dict[str, Any] | None = None,
) -> None:
    """Iterate circular SCC in-place (Gauss-Seidel) until convergence.
    Modifies value_map in place. Silent on non-convergence."""
    # Seed None values with 0 (matches Excel behaviour)
    for _, cell in scc_cells:
        k = id_to_key[cell.stable_id]
        if value_map.get(k) is None:
            value_map[k] = 0.0

    prev_signs: dict[str, int] = {}
    omega = 1.0  # relaxation factor; reduced to 0.5 on detected oscillation

    for _iteration in range(max_iterations):
        max_change = 0.0
        for sheet_name, cell in scc_cells:
            k = id_to_key[cell.stable_id]
            old_val = value_map.get(k)
            new_val = _eval_one_cell(sheet_name, cell, value_map, id_to_key, parser, named_tables)
            if new_val is None:
                continue
            # Apply under-relaxation if oscillation was detected
            if omega < 1.0 and isinstance(old_val, (int, float)):
                new_val = (1.0 - omega) * float(old_val) + omega * float(new_val)
            value_map[k] = new_val  # in-place update (Gauss-Seidel)
            if isinstance(old_val, (int, float)) and isinstance(new_val, (int, float)):
                change = abs(float(new_val) - float(old_val))
                max_change = max(max_change, change)
                # Oscillation: sign of delta flips -> engage damping
                sign = 1 if new_val > old_val else (-1 if new_val < old_val else 0)
                prev = prev_signs.get(cell.stable_id, 0)
                if prev != 0 and sign != 0 and sign != prev:
                    omega = 0.5
                if sign != 0:
                    prev_signs[cell.stable_id] = sign

        if max_change < tolerance:
            break  # converged


def evaluate_patch(
    patches: dict[str, Any],
    workbook: Workbook,
) -> dict[str, Any]:
    """
    Apply {stable_id: new_value} patches and re-evaluate all formula cells.
    Returns {stable_id: computed_value} for every cell in the workbook.
    Uses SCC-based topological evaluation with Gauss-Seidel iteration for
    circular references.
    """
    parser = formulas.Parser()
    named_tables = getattr(workbook, "named_tables", {}) or {}

    # Build value map and bidirectional key mappings
    value_map: dict[str, Any] = {}
    id_to_key: dict[str, str] = {}
    key_to_id: dict[str, str] = {}
    for sheet in workbook.sheets:
        for cell in sheet.cells:
            k = _fkey(sheet.name, cell.row, cell.col)
            value_map[k] = cell.value
            id_to_key[cell.stable_id] = k
            key_to_id[k] = cell.stable_id

    # Apply patches
    for stable_id, new_value in patches.items():
        k = id_to_key.get(stable_id)
        if k:
            value_map[k] = new_value

    # Build dependency graph and cell lookup
    dep_graph = _build_dep_graph(workbook)
    cell_lookup: dict[str, tuple[str, Any]] = {}  # stable_id -> (sheet_name, cell)
    for sheet in workbook.sheets:
        for cell in sheet.cells:
            if cell.formula:
                cell_lookup[cell.stable_id] = (sheet.name, cell)

    # _tarjan_sccs emits SCCs in sources-first topological order (see its
    # docstring). Dependency SCCs are appended before the SCCs that depend on
    # them, so no reversal is needed here.
    sccs = _tarjan_sccs(dep_graph)
    sccs_topo = sccs

    # Evaluate each SCC in topological order
    for scc in sccs_topo:
        formula_nodes = [sid for sid in scc if sid in cell_lookup]
        if not formula_nodes:
            continue  # SCC contains only non-formula (assumption) cells

        is_circular = len(scc) > 1 or scc[0] in dep_graph.get(scc[0], set())

        if is_circular:
            scc_cells = [cell_lookup[sid] for sid in formula_nodes]
            _gauss_seidel(scc_cells, value_map, id_to_key, parser, named_tables=named_tables)
        else:
            # Non-circular singleton: evaluate once
            sid = formula_nodes[0]
            sheet_name, cell = cell_lookup[sid]
            result = _eval_one_cell(sheet_name, cell, value_map, id_to_key, parser, named_tables)
            value_map[id_to_key[sid]] = result

    return {
        key_to_id[k]: _scalar(v)
        for k, v in value_map.items()
        if k in key_to_id
    }
