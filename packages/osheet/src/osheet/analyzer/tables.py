from __future__ import annotations
from openpyxl.utils import get_column_letter
from osheet.models import Cell, Sheet, Table, Column, ColumnDtype


def _cell_grid(cells: list[Cell]) -> dict[tuple[int, int], Cell]:
    return {(c.row, c.col): c for c in cells}


def _is_text(cell: Cell) -> bool:
    return isinstance(cell.value, str) and cell.formula is None


def detect_tables(sheet: Sheet) -> list[Table]:
    if not sheet.cells:
        return []

    grid = _cell_grid(sheet.cells)
    rows = sorted({c.row for c in sheet.cells})
    cols = sorted({c.col for c in sheet.cells})

    if not rows or not cols:
        return []

    min_col, max_col = cols[0], cols[-1]

    # Find header rows: rows where majority of occupied cells are text
    header_candidates: list[int] = []
    for row in rows:
        row_cells = [grid[(row, c)] for c in cols if (row, c) in grid]
        if not row_cells:
            continue
        text_count = sum(1 for c in row_cells if _is_text(c))
        if text_count / len(row_cells) >= 0.5 and len(row_cells) >= 2:
            header_candidates.append(row)

    if not header_candidates:
        header_candidates = [rows[0]]

    tables: list[Table] = []
    for i, header_row in enumerate(header_candidates):
        next_header = header_candidates[i + 1] if i + 1 < len(header_candidates) else None
        data_rows = [r for r in rows if r > header_row and (next_header is None or r < next_header)]
        if not data_rows:
            continue

        first_data = data_rows[0]
        last_data = data_rows[-1]

        header_cols = [c for c in cols if (header_row, c) in grid]
        if not header_cols:
            header_cols = cols

        range_ref = (
            f"{get_column_letter(min(header_cols))}{header_row}:"
            f"{get_column_letter(max(header_cols))}{last_data}"
        )

        columns = []
        for col_idx in header_cols:
            header_cell = grid.get((header_row, col_idx))
            col_name = str(header_cell.value) if header_cell else get_column_letter(col_idx)
            columns.append(Column(name=col_name, dtype=ColumnDtype.UNKNOWN, col_index=col_idx))

        sheet_id = sheet.id.replace("sheet.", "")
        table_id = f"table.{sheet_id}.row{header_row}"

        tables.append(Table(
            id=table_id,
            sheet_name=sheet.name,
            range_ref=range_ref,
            columns=columns,
            header_row=header_row,
            first_data_row=first_data,
            last_data_row=last_data,
            confidence=0.8,
        ))

    return tables
