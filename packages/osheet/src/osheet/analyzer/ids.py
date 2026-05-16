from __future__ import annotations
import re
from openpyxl.utils import get_column_letter
from osheet.models import Cell, CellRole, Sheet, Table, Workbook


def _slug(text: str) -> str:
    """Convert any string to a lowercase dot-safe slug."""
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = text.strip("_")
    return text or "cell"


def _table_for_cell(cell: Cell, tables: list[Table]) -> Table | None:
    for t in tables:
        if cell.row >= t.header_row and cell.row <= t.last_data_row:
            if any(col.col_index == cell.col for col in t.columns):
                return t
    return None


def _col_name_for_cell(cell: Cell, table: Table) -> str | None:
    for col in table.columns:
        if col.col_index == cell.col:
            return col.name
    return None


def assign_stable_ids(workbook: Workbook) -> None:
    """Assign deterministic, human-readable stable IDs to every cell."""
    seen: set[str] = set()

    def unique(base: str) -> str:
        if base not in seen:
            seen.add(base)
            return base
        i = 2
        while f"{base}_{i}" in seen:
            i += 1
        candidate = f"{base}_{i}"
        seen.add(candidate)
        return candidate

    for sheet in workbook.sheets:
        sheet_slug = _slug(sheet.name)
        for cell in sheet.cells:
            table = _table_for_cell(cell, sheet.tables)
            if table:
                table_slug = _slug(table.id.split(".")[-1])
                col_name = _col_name_for_cell(cell, table)
                if col_name and cell.row != table.header_row:
                    col_slug = _slug(col_name)
                    row_label = f"r{cell.row}"
                    if cell.role == CellRole.ASSUMPTION:
                        base = f"assumption.{sheet_slug}.{col_slug}"
                    elif cell.role == CellRole.OUTPUT:
                        base = f"metric.{sheet_slug}.{col_slug}"
                    else:
                        base = f"{sheet_slug}.{table_slug}.{col_slug}.{row_label}"
                else:
                    base = f"{sheet_slug}.{get_column_letter(cell.col)}{cell.row}".lower()
            else:
                col_letter = get_column_letter(cell.col).lower()
                base = f"{sheet_slug}.{col_letter}{cell.row}"

            cell.stable_id = unique(base)
