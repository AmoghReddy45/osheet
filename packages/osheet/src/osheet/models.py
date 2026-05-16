from __future__ import annotations
from enum import Enum
from typing import Any
from pydantic import BaseModel, Field, model_validator


class CellRole(str, Enum):
    ASSUMPTION = "assumption"
    OUTPUT = "output"
    INTERMEDIATE = "intermediate"
    LABEL = "label"
    UNKNOWN = "unknown"


class ColumnDtype(str, Enum):
    NUMBER = "number"
    TEXT = "text"
    DATE = "date"
    FORMULA = "formula"
    MIXED = "mixed"
    UNKNOWN = "unknown"


class Cell(BaseModel):
    stable_id: str
    role: CellRole = CellRole.UNKNOWN
    value: Any = None
    formula: str | None = None
    depends_on: list[str] = Field(default_factory=list)
    sheet_name: str = ""
    col: int = 0
    row: int = 0
    fill_color: str | None = None
    confidence: float = 0.0

    @property
    def address(self) -> str:
        from openpyxl.utils import get_column_letter
        return f"{self.sheet_name}!{get_column_letter(self.col)}{self.row}"


class Column(BaseModel):
    name: str
    dtype: ColumnDtype = ColumnDtype.UNKNOWN
    col_index: int = 0


class Table(BaseModel):
    id: str
    sheet_name: str
    range_ref: str
    columns: list[Column] = Field(default_factory=list)
    header_row: int = 0
    first_data_row: int = 0
    last_data_row: int = 0
    confidence: float = 0.0


class Sheet(BaseModel):
    id: str
    name: str
    tables: list[Table] = Field(default_factory=list)
    cells: list[Cell] = Field(default_factory=list)


class NamedTable(BaseModel):
    name: str
    sheet_name: str
    ref: str                      # e.g. "A1:G25"
    header_row: int               # first row of ref (where column names live)
    first_col: int                # left column index (1-based)
    last_col: int = 0             # right column index (1-based)
    first_data_row: int = 0       # header_row + 1 typically
    last_data_row: int = 0        # bottom of ref minus totals row
    has_totals_row: bool = False
    columns: dict[str, int] = Field(default_factory=dict)  # col_name -> absolute col index (1-based)


class Warning(BaseModel):
    address: str
    message: str
    level: str = "warn"


class Manifest(BaseModel):
    source_file: str = ""
    sheet_count: int = 0
    table_count: int = 0
    assumption_count: int = 0
    output_count: int = 0
    warnings: list[Warning] = Field(default_factory=list)


class Workbook(BaseModel):
    sheets: list[Sheet]
    manifest: Manifest = Field(default_factory=Manifest)
    named_tables: dict[str, NamedTable] = Field(default_factory=dict)

    @property
    def assumptions(self) -> list[Cell]:
        return [c for s in self.sheets for c in s.cells if c.role == CellRole.ASSUMPTION]

    @property
    def outputs(self) -> list[Cell]:
        return [c for s in self.sheets for c in s.cells if c.role == CellRole.OUTPUT]

    @property
    def all_cells(self) -> list[Cell]:
        return [c for s in self.sheets for c in s.cells]

    def get_cell(self, stable_id: str) -> Cell | None:
        return next((c for c in self.all_cells if c.stable_id == stable_id), None)
