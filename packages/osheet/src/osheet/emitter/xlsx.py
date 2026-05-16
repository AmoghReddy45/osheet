from __future__ import annotations
import io
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from osheet.models import CellRole, Workbook


_ROLE_COLORS = {
    CellRole.ASSUMPTION:   "FFD4A96A",
    CellRole.OUTPUT:       "FF6AB07A",
    CellRole.INTERMEDIATE: "FF6A8FD4",
}


def to_xlsx_bytes(workbook: Workbook) -> bytes:
    wb_ox = openpyxl.Workbook()
    wb_ox.remove(wb_ox.active)  # remove default sheet

    for sheet in workbook.sheets:
        ws = wb_ox.create_sheet(title=sheet.name)
        for cell in sheet.cells:
            ox_cell = ws.cell(row=cell.row, column=cell.col)
            ox_cell.value = cell.formula if cell.formula else cell.value

            # Colour-code by role
            color = _ROLE_COLORS.get(cell.role)
            if color:
                ox_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # Embed stable_id as cell comment
            if cell.stable_id:
                comment = Comment(
                    f"stable_id: {cell.stable_id}\nrole: {cell.role.value}\nconf: {cell.confidence:.2f}",
                    "osheet"
                )
                ox_cell.comment = comment

    # Manifest hidden sheet
    meta_ws = wb_ox.create_sheet(title="_ai_manifest")
    meta_ws.sheet_state = "hidden"
    meta_ws["A1"] = "osheet_version"
    meta_ws["B1"] = "0.1.0"
    meta_ws["A2"] = "assumption_count"
    meta_ws["B2"] = workbook.manifest.assumption_count
    meta_ws["A3"] = "output_count"
    meta_ws["B3"] = workbook.manifest.output_count
    meta_ws["A4"] = "table_count"
    meta_ws["B4"] = workbook.manifest.table_count

    buf = io.BytesIO()
    wb_ox.save(buf)
    return buf.getvalue()
