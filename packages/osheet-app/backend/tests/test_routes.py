import io, time
from fastapi.testclient import TestClient
import openpyxl
from app.main import app

client = TestClient(app)


def _make_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Revenue"
    ws["B1"] = 100
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_health():
    r = client.get("/health")
    assert r.status_code == 200


def test_convert_returns_job_id():
    data = _make_xlsx()
    r = client.post("/convert", files={"file": ("model.xlsx", data, "application/octet-stream")})
    assert r.status_code == 200
    assert "job_id" in r.json()


def test_result_after_convert():
    data = _make_xlsx()
    r = client.post("/convert", files={"file": ("model.xlsx", data, "application/octet-stream")})
    job_id = r.json()["job_id"]
    # TestClient with raise_server_exceptions=True runs background tasks synchronously
    r2 = client.get(f"/result/{job_id}")
    assert r2.status_code == 200
    assert r2.json()["status"] in ("done", "processing", "error")


def test_download_xlsx_after_done():
    data = _make_xlsx()
    r = client.post("/convert", files={"file": ("model.xlsx", data, "application/octet-stream")})
    job_id = r.json()["job_id"]
    # Poll until done
    for _ in range(20):
        r2 = client.get(f"/result/{job_id}")
        if r2.json()["status"] == "done":
            break
        time.sleep(0.05)
    r3 = client.get(f"/download/{job_id}/xlsx")
    assert r3.status_code == 200
    assert "application/vnd" in r3.headers["content-type"]


def test_rejects_non_xlsx():
    r = client.post("/convert", files={"file": ("model.csv", b"a,b,c", "text/csv")})
    assert r.status_code == 400


def test_result_not_found():
    r = client.get("/result/nonexistent-job-id")
    assert r.status_code == 404
